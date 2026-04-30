"""
Transform an Excel tab from the input folder to JSON.
Only visible rows are included (hidden rows are skipped).
Extracts the visible (formatted) value from each cell, not the raw stored value.
Output is cleaned: all-null rows and all-null columns removed, null values omitted from each object.
"""

import json
import os
import re
from datetime import date, datetime, time
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Please install openpyxl: pip install openpyxl")
    raise SystemExit(1)

INPUT_DIR = Path(__file__).resolve().parent / "input"

# When set (e.g. "PLN", "CZK", "EUR"), used as currency suffix when the format has no currency or uses "-" as placeholder.
# Set via env: RATE_DEFAULT_CURRENCY=CZK  (e.g. for CZ/SK tariff sheets that show " -" instead of currency)
DEFAULT_CURRENCY = os.environ.get("RATE_DEFAULT_CURRENCY", "").strip()
if DEFAULT_CURRENCY and not DEFAULT_CURRENCY.startswith(" "):
    DEFAULT_CURRENCY = f" {DEFAULT_CURRENCY}"

# Set to True to strip title/fuel/metadata rows from JSON (see output/JSON_CONTENT_ANALYSIS.md).
STRIP_METADATA = True

# Table title row patterns (sheet tables, not tab names). Rows starting a new "table" block.
_TABLE_TITLE_PATTERNS = (
    "Freight Cost",
    "New Freight Rate",
    "FUEL INDEX",
    "Rates RDE",
    "Rates Per Shipment",
    "Return Rates",
    "Rerurn Rates",
)
_MONTHS = ("January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December")


def get_decimal_places(number_format):
    """Infer decimal places from Excel number format (e.g. 0.000 -> 3, 0,00 -> 2)."""
    if not number_format:
        return 2
    # Match pattern like .0 or .00 (decimal part, US style) or ,0 or ,00 (EU style)
    m = re.search(r"[.,](0+)(?![0-9])", number_format)
    if m:
        return len(m.group(1))
    if "#" in number_format or "0" in number_format:
        return 2
    return 2


def _as_currency_suffix(raw: str) -> str:
    """Use the format's currency/suffix as-is; only treat dash-only as no currency."""
    if not raw or not raw.strip():
        return ""
    s = raw.strip()
    # Placeholder-only (dash or only spaces/dashes) -> no currency so DEFAULT_CURRENCY can apply
    if s in ("-", "–", "—") or not s.replace("-", "").replace(" ", "").strip():
        return ""
    return f" {s}"


def get_currency_suffix(number_format):
    """
    Extract currency suffix from Excel number format and use it as-is (no conversion).
    Same text as in the original file: zł stays zł, PLN stays PLN, € stays €, etc.

    1. Quoted text in format (e.g. "PLN", "zł", " -") -> use last quoted string as suffix.
    2. Bracket locale/currency [$CODE] or [$CODE-123] -> use CODE as suffix.
    3. Unquoted € in format -> use " €".
    """
    if not number_format:
        return ""
    nf = number_format.strip()

    # 1. Quoted literal in format
    quoted = re.findall(r'"([^"]*)"', nf)
    if quoted:
        suffix = _as_currency_suffix(quoted[-1])
        if suffix:
            return suffix

    # 2. Bracket currency/locale: [$CODE] or [$CODE-123] or [$"CODE"]
    bracket_match = re.search(r'\[\$["\']?([^"\]\s-]+)["\']?(?:-\d+)?\]', nf)
    if not bracket_match:
        bracket_match = re.search(r"\[\$([^\]-]+)(?:-\d+)?\]", nf)
    if bracket_match:
        code = bracket_match.group(1).strip()
        return _as_currency_suffix(code)

    # 3. Unquoted symbol in format
    if "€" in nf:
        return " €"

    return ""


def format_cell_display(cell):
    """
    Return the value as Excel would display it (formatted), not the raw value.
    E.g. 6.769... with format '0.000 "€"' -> "6.769 €"; with '0.00 "PLN"' -> "6.77 PLN"
    Whole numbers with no currency suffix use integer text (``"1"``) so zones/postals are not ``"1.00"``.
    Always returns JSON-serializable types (str, int, float, None).
    """
    value = cell.value
    if value is None:
        return None
    nf = (cell.number_format or "General").strip()
    if isinstance(value, (int, float)):
        suffix = get_currency_suffix(nf) or (DEFAULT_CURRENCY if DEFAULT_CURRENCY else "")
        if not suffix:
            if isinstance(value, int):
                return str(value)
            if (
                isinstance(value, float)
                and math.isfinite(value)
                and value == int(value)
            ):
                return str(int(value))
        decimals = get_decimal_places(nf)
        try:
            formatted = f"{value:.{decimals}f}"
        except (ValueError, TypeError):
            return value
        if suffix:
            formatted = f"{formatted}{suffix}"
        return formatted
    # Excel dates/times: make JSON-serializable
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    # Fallback: ensure we don't leave non-serializable types (e.g. from openpyxl)
    if isinstance(value, str):
        return value
    return str(value)


def get_xlsx_files():
    """List .xlsx files in the input folder."""
    if not INPUT_DIR.exists():
        print(f"Input folder not found: {INPUT_DIR}")
        return []
    files = sorted(f.name for f in INPUT_DIR.iterdir() if f.suffix.lower() == ".xlsx")
    return files


def is_row_hidden(ws, row_num):
    """Return True if the row is hidden."""
    if row_num not in ws.row_dimensions:
        return False
    return bool(ws.row_dimensions[row_num].hidden)


def sheet_to_json(ws):
    """
    Convert worksheet to list of dicts. All columns are Column0, Column1, Column2, ...
    so that cell text like "current share" or "diesel" stays as values, not as key names.
    Skips hidden rows. Uses each cell's displayed (formatted) value, not raw value.
    """
    rows = []
    for row_num, row in enumerate(ws.iter_rows(), start=1):
        if is_row_hidden(ws, row_num):
            continue
        row_dict = {f"Column{i}": format_cell_display(cell) for i, cell in enumerate(row)}
        rows.append(row_dict)
    return rows


def _is_empty_value(v):
    return v is None or (isinstance(v, str) and v.strip() == "")


# Substrings that mark a row as metadata (not tariff data). Table names like "Rates RDE2...", "New Freight Rate..." are kept.
_METADATA_MARKERS = (
    "diesel share",
    "Fuel INDEX",
    "Difference",
    "over 5% Change",
    "Freight Cost without Fuel",
    "decrease freight",
)
_ISO_DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}")


def _row_looks_like_tariff_measure_banner_row(row: dict) -> bool:
    """
    Row that labels Flat / per-kg band columns (e.g. Price per shipment, Price per kg).
    Must **not** be classified as fuel/calculation metadata — including after dual-title expansion,
    when the same row is reindexed to **two** cells and would otherwise match the small-row
    ``Price per shipment`` + ``Price per kg`` strip rule.
    """
    blob = " ".join(str(v) for v in row.values() if v is not None and str(v).strip()).lower()
    if not blob:
        return False
    if "price per" in blob and ("shipment" in blob or "kg" in blob):
        return True
    if "rates per" in blob and "shipment" in blob:
        return True
    if re.search(r"rate\s*per\s*kg", blob):
        return True
    if re.search(r"per\s*[\d.,]+\s*kg", blob):
        return True
    return False


def _is_metadata_row(row: dict) -> bool:
    """True if this row is fuel/calculation metadata only. Table names (e.g. Rates RDE2..., New Freight Rate...) are kept."""
    if _row_looks_like_tariff_measure_banner_row(row):
        return False
    vals = [str(v) for v in row.values() if v is not None and str(v).strip()]
    if not vals:
        return False
    combined = " ".join(vals).lower()
    for marker in _METADATA_MARKERS:
        if marker.lower() in combined:
            return True
    if len(row) <= 4 and any(_ISO_DATE_RE.match(str(v)) for v in row.values() if v):
        return True
    # Drop tiny junk rows that only repeat price-tier labels with no table context. Do **not** drop
    # wide dual-lane tier banners (e.g. FIEGE: Column2/9 + Column29/36 = two "Price per shipment" +
    # two "Price per kg") — those are 4 populated cells and must stay for Flat / p/unit mapping.
    if (
        len(row) <= 4
        and len(vals) <= 3
        and "Price per shipment" in vals
        and "Price per kg" in vals
    ):
        if not any(z in combined for z in ("zip", "area code", "destination postal code")):
            return True
    return False


# Stray Excel cells: e.g. {"Column12": "0.02"} with no postal / lane context (rest of row empty).
_POSTAL_OR_ZONE_IN_C0 = re.compile(
    r"\d{3}\s+\d{2}\s*[-–]\s*\d{3}\s+\d{2}|^\d+\s*[-–]\s*\d+\s*$|zip|postal|destination|area\s*code",
    re.I,
)


def _cell_looks_like_standalone_rate_amount(s: str) -> bool:
    """True if the cell is only a number, optionally with a currency suffix (tariff amount)."""
    s = str(s).strip().replace("\u00a0", " ")
    if not s or len(s) > 80:
        return False
    t = re.sub(
        r"\s*(Kč|€|EUR|USD|GBP|CHF|PLN|CZK|HUF|zł|zl)\s*$",
        "",
        s,
        flags=re.I,
    ).strip()
    t = re.sub(r"\s+", "", t)
    if not t:
        return False
    if t.count(",") == 1 and t.count(".") == 0:
        parts = t.split(",")
        if len(parts) == 2 and len(parts[1]) <= 2 and parts[0].replace("-", "").isdigit():
            t = parts[0] + "." + parts[1]
        elif len(parts) == 2:
            t = t.replace(",", ".")
    elif t.count(",") > 1:
        return False
    return bool(re.match(r"^-?\d+(\.\d+)?$", t))


def _column_key_sort(key: str) -> int:
    m = re.match(r"^Column(\d+)$", str(key))
    return int(m.group(1)) if m else 10**9


def _row_has_tariff_lane_context(row: dict) -> bool:
    """True if the leftmost populated cell looks like a zone / header / lane, not a lone amount."""
    for key in sorted(row.keys(), key=_column_key_sort):
        if not str(key).startswith("Column"):
            continue
        v = row.get(key)
        if _is_empty_value(v):
            continue
        s = str(v).strip()
        if _POSTAL_OR_ZONE_IN_C0.search(s):
            return True
        if len(s) > 25 and re.search(r"[A-Za-z]{6,}", s):
            return True
        break
    return False


def _is_sparse_junk_numeric_row(row: dict) -> bool:
    """
    Drop rows that are effectively one/few stray numeric cells (e.g. misaligned '0.02')
    with no postal band or label in the leading columns.
    """
    if _row_looks_like_tariff_measure_banner_row(row):
        return False
    items = [(k, v) for k, v in row.items() if not _is_empty_value(v)]
    if not items or len(items) > 5:
        return False
    if _row_has_tariff_lane_context(row):
        return False
    for _k, v in items:
        s = str(v).strip()
        if len(s) > 120:
            return False
        if len(s) > 12 and not _cell_looks_like_standalone_rate_amount(s):
            return False
    for _k, v in items:
        if not _cell_looks_like_standalone_rate_amount(str(v)):
            return False
    return True


def _column_key_index(ck: str) -> int:
    """Numeric index from ``Column12`` → 12."""
    m = re.match(r"^Column(\d+)$", str(ck))
    return int(m.group(1)) if m else -1


def _cells_matching_table_title_patterns(row: dict) -> list[tuple[str, str]]:
    """
    (column_key, text) for each cell that matches a :data:`_TABLE_TITLE_PATTERNS` substring
    and is long enough for :func:`_get_table_title_from_row` (>= 15 chars).
    Left-to-right order.
    """
    out: list[tuple[str, str]] = []
    for ck in sorted(row.keys(), key=_column_key_sort):
        v = row.get(ck)
        if v is None or not str(v).strip():
            continue
        s = str(v).strip()
        if len(s) < 15:
            continue
        for pat in _TABLE_TITLE_PATTERNS:
            if pat in s:
                out.append((ck, s))
                break
    return out


def _find_wide_matrix_header_row_index(rows: list[dict], dual_idx: int, *, max_scan: int = 8) -> int | None:
    """
    Row index of the first wide header that has two lane starts (second Zip 2 / Currency /
    Destination Postal Code), scanning **forward** from ``dual_idx + 1``.

    Some sheets insert a **tier** row (e.g. Price per shipment / Price per kg) between the
    dual-title row and the band header; the lane split must not assume ``dual_idx + 1`` is the header.
    """
    end = min(len(rows), dual_idx + 1 + max_scan)
    for j in range(dual_idx + 1, end):
        if _find_second_lane_header_split_ck(rows[j]):
            return j
    return None


def _find_second_lane_header_split_ck(header_row: dict) -> str | None:
    """
    Column key where the **second** lane starts in a wide side-by-side header row:
    second ``Zip 2``, second ``Currency``, or second ``Destination Postal Code``-like label.
    """
    keys = sorted(header_row.keys(), key=_column_key_sort)
    zip2_cks: list[str] = []
    currency_cks: list[str] = []
    dest_cks: list[str] = []
    for ck in keys:
        v = header_row.get(ck)
        if v is None or not str(v).strip():
            continue
        t = str(v).strip().lower()
        if t in ("zip 2", "zip2") or re.match(r"^zip\s*2$", t):
            zip2_cks.append(ck)
        elif t == "currency":
            currency_cks.append(ck)
        elif "destination postal" in t:
            dest_cks.append(ck)
    if len(zip2_cks) >= 2:
        return zip2_cks[1]
    if len(currency_cks) >= 2:
        return currency_cks[1]
    if len(dest_cks) >= 2:
        return dest_cks[1]
    return None


def _project_row_reindex_columns(row: dict, column_keys: list[str]) -> dict:
    """Keep only given keys, renumber to ``Column0`` … ``Column{n-1}``, omit empties."""
    out: dict = {}
    for i, ck in enumerate(column_keys):
        v = row.get(ck)
        if v is not None and not _is_empty_value(v):
            out[f"Column{i}"] = v
    return out


def _expand_dual_title_row_side_by_side(rows: list[dict]) -> list[dict]:
    """
    Some sheets put **two** table titles on **one** row (e.g. ``Rates RDE1…`` in Column1 and
    ``Return Rates RDE1…`` in Column28), with a wide header row below (two ``Destination Postal Code``
    blocks). That yields **one** JSON row for titles and merged columns — downstream expects **two**
    stacked tables (title → matrix → title → matrix), like
    ``Bridgestone Distribution tariff - February CZ_Tariff sheet Bor CZ - February.json``.

    When we see a row with **two** table-title cells, we find the wide **band header** row (two lane
    starts) by scanning a few rows down — a **tier** row (Price per shipment / per kg) may sit
    between title and header. Then split into: left title + left columns for all following rows,
    then right title + right columns.

    Disable with ``RATE_DISABLE_DUAL_TITLE_SPLIT=1``.
    """
    if not rows:
        return rows
    if os.environ.get("RATE_DISABLE_DUAL_TITLE_SPLIT", "").strip().lower() in (
        "1",
        "true",
        "yes",
        "y",
    ):
        return rows

    dual_idx: int | None = None
    for i, row in enumerate(rows):
        if len(_cells_matching_table_title_patterns(row)) >= 2:
            dual_idx = i
            break
    if dual_idx is None or dual_idx + 1 >= len(rows):
        return rows

    hi = _find_wide_matrix_header_row_index(rows, dual_idx)
    if hi is None:
        return rows
    split_ck = _find_second_lane_header_split_ck(rows[hi])
    if not split_ck:
        return rows
    split_n = _column_key_index(split_ck)
    if split_n < 0:
        return rows

    all_cks: set[str] = set()
    for r in rows[dual_idx + 1 :]:
        all_cks.update(r.keys())
    sorted_cks = sorted(all_cks, key=_column_key_sort)
    left_cks = [ck for ck in sorted_cks if _column_key_index(ck) < split_n]
    right_cks = [ck for ck in sorted_cks if _column_key_index(ck) >= split_n]
    if len(left_cks) < 2 or len(right_cks) < 2:
        return rows

    title_pairs = _cells_matching_table_title_patterns(rows[dual_idx])
    title_left = title_pairs[0][1].strip()
    title_right = title_pairs[1][1].strip()

    out: list[dict] = []
    out.extend(rows[:dual_idx])
    # Table 1: origin / outbound lane (single-cell title rows — avoid raw leading spaces from Excel)
    out.append({"Column0": title_left})
    for r in rows[dual_idx + 1 :]:
        out.append(_project_row_reindex_columns(r, left_cks))
    # Table 2: return lane
    out.append({"Column0": title_right})
    for r in rows[dual_idx + 1 :]:
        out.append(_project_row_reindex_columns(r, right_cks))
    return out


def _clean_data(data: list[dict]) -> list[dict]:
    """Remove all-null rows, all-null columns, and omit null values from each object. Optionally strip metadata rows."""
    if not data:
        return data
    all_keys = set()
    for row in data:
        all_keys.update(row.keys())
    non_empty_rows = [row for row in data if not all(_is_empty_value(row.get(k)) for k in all_keys)]
    data = non_empty_rows
    if not data:
        return data
    if STRIP_METADATA:
        data = [row for row in data if not _is_metadata_row(row)]
    if not data:
        return data
    data = [row for row in data if not _is_sparse_junk_numeric_row(row)]
    if not data:
        return data
    keys_with_values = set()
    for row in data:
        for k, v in row.items():
            if not _is_empty_value(v):
                keys_with_values.add(k)
    keep_keys = keys_with_values if keys_with_values else all_keys
    cleaned = []
    for row in data:
        obj = {k: v for k, v in row.items() if k in keep_keys and not _is_empty_value(v)}
        if obj:
            cleaned.append(obj)
    cleaned = _expand_dual_title_row_side_by_side(cleaned)
    return cleaned


def clean_tariff_rows_twice(data: list[dict]) -> list[dict]:
    """
    Apply :func:`_clean_data` twice — same net effect as exporting with ``main()`` (one pass) and
    then running ``clean_rate_json.py`` on the file (second pass).
    """
    return _clean_data(_clean_data(data))


def export_tariff_sheet_rows(path: Path, sheet_name: str) -> list[dict]:
    """
    One workbook tab through the same pipeline as ``main()``: ``sheet_to_json`` →
    ``_extract_tables_by_name`` (uses ``path.stem`` for keyword table selection) →
    ``clean_tariff_rows_twice``.
    """
    wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
    ws = wb[sheet_name]
    data = sheet_to_json(ws)
    data = _extract_tables_by_name(data, path.stem, sheet_name)
    return clean_tariff_rows_twice(data)


def combine_tariff_sheet_rows(file_sheet_pairs: list[tuple[Path, str]]) -> list[dict]:
    """
    Concatenate tariff rows from several ``(xlsx_path, sheet_name)`` sources in order.
    Each file is processed independently (correct stem for ``_extract_tables_by_name``), then rows
    are appended so the result is one JSON array with the same row shape as a single export.
    """
    out: list[dict] = []
    for path, sheet_name in file_sheet_pairs:
        out.extend(export_tariff_sheet_rows(path, sheet_name))
    return out


def combined_tariff_output_basename(file_sheet_pairs: list[tuple[Path, str]]) -> str:
    """
    Filename stem for merged tariff JSON (no extension), e.g.
    ``combined_tariff_2files__Sheet1`` or ``combined_tariff_2files__Sheet1+Tariff``.
    """
    if not file_sheet_pairs:
        return "combined_tariff_empty"
    sheets = [sn for _, sn in file_sheet_pairs]
    uniq = list(dict.fromkeys(sheets))
    n = len(file_sheet_pairs)

    def _safe(s: str) -> str:
        return re.sub(r'[<>:"/\\|?*\n\r\t]', "_", s).strip()[:80]

    if len(uniq) == 1:
        return f"combined_tariff_{n}files__{_safe(uniq[0])}"
    joined = "+".join(_safe(s) for s in uniq)[:120]
    return f"combined_tariff_{n}files__{joined}"


def parse_tariff_file_index_list(choice: str, n_files: int) -> list[int]:
    """
    Parse user input like ``1``, ``1,3``, ``1, 2, 4`` or ``1-3`` into unique 1-based indices
    in order (each between 1 and ``n_files``).
    """
    choice = (choice or "").strip()
    if not choice:
        return []
    indices: list[int] = []
    for part in re.split(r"[,\s;]+", choice):
        part = part.strip()
        if not part:
            continue
        m = re.match(r"^(\d+)\s*-\s*(\d+)$", part.strip())
        if m:
            lo, hi = int(m.group(1)), int(m.group(2))
            if lo > hi:
                lo, hi = hi, lo
            for k in range(lo, hi + 1):
                indices.append(k)
        else:
            indices.append(int(part))
    seen: set[int] = set()
    out: list[int] = []
    for i in indices:
        if 1 <= i <= n_files and i not in seen:
            seen.add(i)
            out.append(i)
    return out


def _get_table_title_from_row(row: dict) -> str | None:
    """If row looks like a table title row, return the title string; else None."""
    candidates = [str(v).strip() for v in row.values() if v is not None and str(v).strip()]
    if not candidates:
        return None
    for s in sorted(candidates, key=len, reverse=True):
        if len(s) < 15:
            continue
        for pat in _TABLE_TITLE_PATTERNS:
            if pat in s:
                return s
    return None


def _split_into_tables(rows: list[dict]) -> list[list[dict]]:
    """Split rows into table blocks. Each block starts with a table title row."""
    tables = []
    current = []
    for row in rows:
        title = _get_table_title_from_row(row)
        if title:
            if current:
                tables.append(current)
            current = [row]
        else:
            current.append(row)
    if current:
        tables.append(current)
    return tables


def _extract_keywords_from_name(file_stem: str, sheet_name: str) -> list[str]:
    """Extract month names and 2-letter tokens (e.g. SK, CZ) from file name and tab name for table matching."""
    text = f"{file_stem} {sheet_name}"
    words = re.findall(r"[A-Za-z]+", text)
    keywords = []
    for w in words:
        if w in _MONTHS:
            keywords.append(w)
        elif len(w) == 2 and w.isalpha():
            keywords.append(w.upper())
    return list(dict.fromkeys(keywords))


def _select_tables(tables: list[list[dict]], keywords: list[str]) -> list[list[dict]]:
    """
    Keep tables whose title contains keywords (from filename/tab); if none match, keep last 2 tables.
    If several tables tie for the best score, keep only the **last two** (sheet order), e.g. outbound
    + returns after older blocks like tolls.
    """
    if not tables:
        return []
    if len(tables) <= 2:
        return tables
    if not keywords:
        return tables[-2:]
    scored = []
    for t in tables:
        title = _get_table_title_from_row(t[0]) if t else ""
        if not title:
            scored.append((0, t))
            continue
        title_upper = title.upper()
        score = sum(1 for k in keywords if k.upper() in title_upper)
        scored.append((score, t))
    best = max(s[0] for s in scored)
    if best > 0:
        sel = [t for score, t in scored if score == best]
        if len(sel) > 2:
            return sel[-2:]
        return sel
    return tables[-2:]


def _extract_tables_by_name(rows: list[dict], file_stem: str, sheet_name: str) -> list[dict]:
    """If sheet has multiple tables, keep only those matching filename/tab keywords, else last 2. Flatten to rows."""
    tables = _split_into_tables(rows)
    if len(tables) <= 1:
        return rows
    keywords = _extract_keywords_from_name(file_stem, sheet_name)
    selected = _select_tables(tables, keywords)
    return [row for block in selected for row in block]


def main():
    files = get_xlsx_files()
    if not files:
        print("No .xlsx files found in the input folder.")
        return

    print("Files in 'input' folder:")
    for i, name in enumerate(files, 1):
        print(f"  {i}. {name}")
    choice = input(
        "Which file number(s) to transform? One number, or several separated by commas "
        "(e.g. 1,3 or 1-3): "
    ).strip()
    indices = parse_tariff_file_index_list(choice, len(files))
    if not indices:
        print("No valid file numbers.")
        return

    pairs: list[tuple[Path, str]] = []
    for fi in indices:
        filename = files[fi - 1]
        path = INPUT_DIR / filename
        wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
        sheet_names = wb.sheetnames
        print(f"\n--- {filename} ---\nTabs:")
        for i, name in enumerate(sheet_names, 1):
            print(f"  {i}. {name}")
        tab_choice = input(f"Tab number for this file [{filename}]: ").strip()
        try:
            tab_idx = int(tab_choice)
            if tab_idx < 1 or tab_idx > len(sheet_names):
                raise ValueError("Invalid number")
            sheet_name = sheet_names[tab_idx - 1]
        except (ValueError, IndexError):
            print("Invalid tab — aborting.")
            return
        pairs.append((path, sheet_name))

    if len(pairs) == 1:
        data = export_tariff_sheet_rows(pairs[0][0], pairs[0][1])
        out_name = f"{pairs[0][0].stem}_{pairs[0][1]}.json"
    else:
        data = combine_tariff_sheet_rows(pairs)
        out_name = combined_tariff_output_basename(pairs) + ".json"
        print("\nMerged sources (in order):")
        for p, sn in pairs:
            print(f"  • {p.name}  →  tab {sn!r}")

    out_path = Path(__file__).resolve().parent / "output" / out_name
    out_path.parent.mkdir(parents=True, exist_ok=True)

    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

    print(f"\nSaved {len(data)} rows to: {out_path}")


if __name__ == "__main__":
    main()
