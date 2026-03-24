"""
Extract rate card information from Excel to JSON:
- General info tab: Rate ID, Validity period
- Rate card tab: Conditional rules (with column tag), Costs (name, rules, rates table)
"""
from __future__ import annotations

import json
import re
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Please install openpyxl: pip install openpyxl")
    raise SystemExit(1)

INPUT_DIR = Path(__file__).resolve().parent / "previous rate"
OUTPUT_DIR = Path(__file__).resolve().parent / "output"

# Expected sheet names (case-insensitive match)
GENERAL_INFO_SHEET = "general info"
RATE_CARD_SHEET = "rate card"


def _cell_value(cell_or_value):
    """Get value as string; accept openpyxl Cell or raw value. Format dates, avoid None."""
    if cell_or_value is None:
        return None
    v = cell_or_value.value if hasattr(cell_or_value, "value") else cell_or_value
    if v is None:
        return None
    if hasattr(v, "isoformat"):
        return v.isoformat()
    return str(v).strip() if isinstance(v, str) else v


def _find_sheet(wb, name: str):
    """Return sheet whose name contains the given name (case-insensitive) or None."""
    key = name.strip().lower()
    for s in wb.sheetnames:
        if key in s.strip().lower():
            return wb[s]
    return None


def extract_general_info(ws) -> dict:
    """Extract Rate ID and Validity period from General info tab."""
    result = {"rate_id": None, "validity_period": None}
    if ws is None:
        return result
    for row in ws.iter_rows(values_only=True):
        for i, val in enumerate(row):
            if val is None:
                continue
            s = str(val).strip().lower()
            if "rate id" in s or s == "rate id":
                next_val = row[i + 1] if i + 1 < len(row) else None
                result["rate_id"] = _cell_value(next_val) if next_val is not None else None
                if result["rate_id"] is not None and not isinstance(result["rate_id"], str):
                    result["rate_id"] = str(result["rate_id"])
            if "validity period" in s or s == "validity period":
                next_val = row[i + 1] if i + 1 < len(row) else None
                result["validity_period"] = _cell_value(next_val) if next_val is not None else None
                if result["validity_period"] is not None and not isinstance(result["validity_period"], str):
                    result["validity_period"] = str(result["validity_period"])
    return result


def _get_cell_value(ws, row_idx, col_idx):
    """Get value at 1-based row, 1-based column."""
    row = row_idx
    col = col_idx
    c = ws.cell(row=row, column=col)
    return _cell_value(c)


# Pattern: "1. 6000 - 6899: FROMPOSTALCODE starts with ..." or "1. CH / LI: FROMCOUNTRY equals CH,LI"
_RULE_LINE_PATTERN = re.compile(r"^\s*\d+\.\s+", re.IGNORECASE)


def extract_conditional_rules(ws) -> list:
    """
    Find the "Conditional rules:" section and parse numbered rule lines.
    Tag_to_apply = the cell directly **below** the rule block in the same column
    (in Excel the conditional rule is above the field to apply).
    """
    rules = []
    if ws is None:
        return rules
    max_row = ws.max_row
    max_col = ws.max_column or 50
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            val = _get_cell_value(ws, r, c)
            if not val:
                continue
            s = str(val).strip()
            if "conditional rules:" not in s.lower():
                continue
            # Found "Conditional rules:" in this column – collect rule lines and find last row of the block
            block_rules = []
            last_rule_row = r
            lines = re.split(r"[\r\n]+", s)
            for line in lines:
                line = line.strip()
                if not line or line.lower().startswith("conditional rules:"):
                    continue
                if _RULE_LINE_PATTERN.match(line) and ":" in line:
                    block_rules.append(line)
            # Check next rows in same column for more rule lines
            for dr in range(1, 30):
                nr = r + dr
                if nr > max_row:
                    break
                val_next = _get_cell_value(ws, nr, c)
                if not val_next:
                    break
                line = str(val_next).strip()
                if not line or "conditional rules" in line.lower():
                    break
                if _RULE_LINE_PATTERN.match(line) and ":" in line:
                    block_rules.append(line)
                    last_rule_row = nr
                else:
                    break
            # Tag_to_apply = cell directly below the rule block (same column)
            tag_row = last_rule_row + 1
            tag = None
            if tag_row <= max_row:
                tag = _get_cell_value(ws, tag_row, c)
                tag = str(tag).strip() if tag else None
            if not tag:
                tag = f"Column{c}"
            for line in block_rules:
                rules.append({"Conditional rule": line, "Tag_to_apply": tag})
    # Dedupe
    seen = set()
    out = []
    for r in rules:
        key = (r["Conditional rule"], r["Tag_to_apply"])
        if key not in seen:
            seen.add(key)
            out.append(r)
    return out


# Horizontal width (columns) for one Transport cost block — avoids joining Applies if from other costs
# when multiple costs sit on the same row (Advanced Export layout).
_COST_BLOCK_COL_SPAN = 24


def _row_values_in_cost_band(ws, row_idx: int, col_start: int, max_col: int) -> list:
    """Cell values only from col_start .. col_start+span-1 (one cost block strip)."""
    c_end = min(col_start + _COST_BLOCK_COL_SPAN - 1, max_col)
    return [_get_cell_value(ws, row_idx, col) for col in range(col_start, c_end + 1)]


def _row_join_band(row_vals: list) -> str:
    return " ".join(str(v) for v in row_vals if v is not None and str(v).strip())


def _extract_applies_if_for_cost_band(
    ws,
    cost_title_row: int,
    col_start: int,
    max_row: int,
    max_col: int,
    *,
    stop_before_row: int | None = None,
) -> list[str]:
    """
    Collect Applies if text only from the same column band as this Transport cost (not the whole sheet row).
    Stops before the lane/rates table row when ``stop_before_row`` is set, or when a lane header is seen in-band.
    """
    c_end = min(col_start + _COST_BLOCK_COL_SPAN - 1, max_col)
    last_nr = min(cost_title_row + 30, max_row)
    if stop_before_row is not None:
        last_nr = min(last_nr, stop_before_row - 1)
    chunks: list[str] = []
    for nr in range(cost_title_row + 1, last_nr + 1):
        cells: list[str] = []
        for col in range(col_start, c_end + 1):
            v = _get_cell_value(ws, nr, col)
            if v is not None and str(v).strip():
                cells.append(str(v).strip())
        if not cells:
            continue
        joined = " ".join(cells)
        low = joined.lower()
        if "lane" in low and ("carrier" in low or "origin" in low):
            break
        if "applies if" in low and "load" in low:
            chunks.append(joined)
        elif chunks and re.match(r"^\d+\.\s", joined):
            # continuation of numbered rule in next row (same block)
            chunks.append(joined)
    if not chunks:
        return []
    text = "\n".join(chunks).strip()
    # Single cell sometimes repeats "Applies if:" — keep first block only
    if text.lower().count("applies if") > 1:
        parts = re.split(r"(?i)(?=\bApplies if\b)", text)
        parts = [p.strip() for p in parts if p.strip()]
        if parts:
            text = parts[0].strip()
            if not text.lower().startswith("applies if"):
                text = "Applies if: " + text
    return [text] if text else []


def _pick_applies_if_line_for_cost_name(cost_name: str, text: str) -> str:
    """
    If text still contains multiple tier rules in one string, pick the line that matches this cost
    (e.g. 'above 20000 kg' -> line with greater than '20000'; plain 'Transport cost' -> line without greater than).
    """
    name = (cost_name or "").strip()
    lines = [ln.strip() for ln in re.split(r"[\r\n]+", text) if ln.strip()]
    one_line = " ".join(lines) if len(lines) <= 1 else text

    m_above = re.search(r"above\s+([\d\s]+)\s*kg", name, re.I)
    if m_above:
        raw_t = re.sub(r"\s+", "", m_above.group(1))
        for ln in lines:
            if "greater than" in ln.lower() and raw_t in re.sub(r"\s+", "", ln):
                return ln
        for ln in lines:
            if f"'{raw_t}'" in ln or f'"{raw_t}"' in ln or raw_t in ln:
                return ln

    if re.fullmatch(r"Transport cost", name, re.I):
        for ln in lines:
            lo = ln.lower()
            if "load" in lo and "greater than" not in lo:
                return ln

    return one_line if len(lines) <= 1 else text


def extract_costs_and_rates_table(ws) -> tuple:
    """
    Find cost blocks and the rates table. Returns (costs, rates_table).
    Costs: name, rate_by, "Calculation rule" (e.g. "Direct rule"), applies_if[].
    Each cost has its own Applies if, read only from that cost's column band (not merged across the row).
    Rates table: separate block, columns before first "Currency".
    """
    costs = []
    rates_table = []
    if ws is None:
        return costs, rates_table
    transport_cost_pattern = re.compile(r"Transport cost(\s*\([^)]+\))?", re.IGNORECASE)
    max_row = ws.max_row
    max_col = ws.max_column or 50
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            val = _get_cell_value(ws, r, c)
            if not val:
                continue
            s = str(val).strip()
            if not transport_cost_pattern.match(s):
                continue
            cost_name = s
            rate_by = None
            calculation_rule = None
            table_header_row = None
            table_data_start = None
            for dr in range(1, 20):
                nr = r + dr
                if nr > max_row:
                    break
                row_vals = _row_values_in_cost_band(ws, nr, c, max_col)
                row_str = _row_join_band(row_vals)
                row_full = [_get_cell_value(ws, nr, col) for col in range(1, max_col + 1)]
                row_str_full = " ".join(str(v) for v in row_full if v is not None and str(v).strip())
                if "rate by" in row_str.lower():
                    for i, v in enumerate(row_vals):
                        if v and "rate by" in str(v).lower():
                            if i + 1 < len(row_vals) and row_vals[i + 1]:
                                rate_by = str(row_vals[i + 1]).strip()
                            break
                # Calculation rule: short phrase only (Direct rule, Minimum rule), not "Conditional rules: ..."
                if re.search(r"\b(Direct rule|Minimum rule)\b", row_str, re.IGNORECASE):
                    calculation_rule = re.search(r"(Direct rule|Minimum rule)", row_str, re.IGNORECASE).group(1)
                elif "rule" in row_str.lower() and "conditional" not in row_str.lower() and calculation_rule is None and len(row_str) < 80:
                    calculation_rule = row_str
                # Lane header can sit outside this cost band — use full row
                if "lane" in row_str_full.lower() and (
                    "carrier" in row_str_full.lower() or "origin" in row_str_full.lower()
                ):
                    table_header_row = nr
                    table_data_start = nr + 1
                    break
            applies_if_raw = _extract_applies_if_for_cost_band(
                ws,
                r,
                c,
                max_row,
                max_col,
                stop_before_row=table_header_row,
            )
            applies_if: list[str] = []
            for blob in applies_if_raw:
                picked = _pick_applies_if_line_for_cost_name(cost_name, blob)
                if picked:
                    applies_if.append(picked)
            if table_header_row is None:
                table_data_start = r + 5
                table_header_row = r + 4
            headers = [_get_cell_value(ws, table_header_row, col) for col in range(1, max_col + 1)]
            headers = [str(h).strip() if h else f"Column{i}" for i, h in enumerate(headers)]
            currency_idx = next((i for i, h in enumerate(headers) if str(h).strip().lower() == "currency"), len(headers))
            headers = headers[:currency_idx]
            table_rows = []
            # Read until sheet end or first fully empty row — do not cap at 200 rows; large rate
            # cards (e.g. Advanced Export) often have 300+ lanes (1, 1-RETURN, 1/4, 1-RETURN/4-RETURN…).
            for tr in range(table_data_start, max_row + 1):
                row_dict = {}
                has_any = False
                for col, h in enumerate(headers):
                    if col + 1 > max_col:
                        break
                    v = _get_cell_value(ws, tr, col + 1)
                    if v is not None and str(v).strip():
                        has_any = True
                    key = h or f"Column{col}"
                    row_dict[key] = v
                if not has_any:
                    break
                table_rows.append(row_dict)
            costs.append({
                "name": cost_name,
                "rate_by": rate_by or "Weight/kg",
                "Calculation rule": calculation_rule,
                "applies_if": applies_if,
            })
            if not rates_table and table_rows:
                rates_table = table_rows
    return costs, rates_table


# Max distinct values for a key to be considered "stable" (lane-defining); keys with more are "changing"
_STABLE_THRESHOLD_MAX = 25
_STABLE_FRACTION = 0.02  # stable if distinct_count <= num_rows * this
# Never treat these as pattern keys (always "changing")
_PATTERN_KEY_EXCLUDE = {"Lane #"}
# Postals may be "changing" globally (many bands) but constant within a pattern — enrich pattern dicts
_POSTAL_KEYS_GROUP_ENRICH = ("Origin Postal Code", "Destination Postal Code")


def _row_matches_pattern_on_stable(row: dict, pattern: dict, stable_keys: list[str]) -> bool:
    for k in stable_keys:
        a = str(row.get(k, "") or "").strip()
        b = str(pattern.get(k, "") or "").strip()
        if a != b:
            return False
    return True


def enrich_patterns_with_group_constant_postals(
    patterns: list[dict],
    stable_keys: list[str],
    rates_table: list,
) -> list[dict]:
    """
    For each pattern, consider all rates_table rows that match it on ``stable_keys``.
    If every such row shares the same non-empty Origin (or Destination) postal code, add that
    field to the pattern — e.g. hub ``34802`` on all outbound lanes and on all return lanes,
    even when the same column varies across the full table (ranges on the other pattern).
    """
    if not patterns or not rates_table or not stable_keys:
        return patterns
    rows = [r for r in rates_table if isinstance(r, dict)]
    out: list[dict] = []
    for p in patterns:
        merged = dict(p)
        matching = [r for r in rows if _row_matches_pattern_on_stable(r, merged, stable_keys)]
        for pk in _POSTAL_KEYS_GROUP_ENRICH:
            vals: set[str] = set()
            for r in matching:
                v = r.get(pk)
                s = str(v).strip() if v is not None and str(v).strip() != "" else ""
                if s:
                    vals.add(s)
            if len(vals) == 1:
                merged[pk] = next(iter(vals))
        out.append(merged)
    return out


def _infer_pattern_keys(rates_table: list) -> tuple[list[str], list[str], dict[str, int]]:
    """
    Infer which keys are stable (same across many rows, define the lane) vs changing (vary from row to row).
    Returns (stable_keys, changing_keys, distinct_count_by_key).
    """
    if not rates_table:
        return [], [], {}
    rows = [r for r in rates_table if isinstance(r, dict) and r]
    if not rows:
        return [], [], {}
    all_keys = []
    seen_keys = set()
    for r in rows:
        for k in r:
            if k not in seen_keys:
                seen_keys.add(k)
                all_keys.append(k)
    num_rows = len(rows)
    threshold = min(_STABLE_THRESHOLD_MAX, max(2, int(num_rows * _STABLE_FRACTION)))
    distinct_count = {}
    for k in all_keys:
        vals = set()
        for r in rows:
            v = r.get(k)
            v = str(v).strip() if v is not None and str(v).strip() != "" else ""
            vals.add(v)
        distinct_count[k] = len(vals)
    stable_keys = [k for k in all_keys if k not in _PATTERN_KEY_EXCLUDE and distinct_count[k] <= threshold]
    changing_keys = [k for k in all_keys if k in _PATTERN_KEY_EXCLUDE or distinct_count[k] > threshold]
    return stable_keys, changing_keys, distinct_count


def extract_patterns(rates_table: list) -> tuple[list, list[str], list[str], dict[str, int]]:
    """
    From rates_table, infer pattern keys from the data: keys that change little (stable) define
    the lane; keys that change a lot are varying dimensions. Returns (patterns, stable_keys,
    changing_keys, distinct_count_by_key).

    Each pattern dict is then enriched with Origin / Destination postal codes when that value is
    the same for every row matching the pattern (e.g. hub ``34802`` on all outbound rows even if
    the other pattern uses ranges in that column). ``changing_keys`` is unchanged (still reflects
    the full table).
    """
    if not rates_table:
        return [], [], [], {}
    stable_keys, changing_keys, distinct_count = _infer_pattern_keys(rates_table)
    if not stable_keys:
        return [], [], list(distinct_count), distinct_count
    seen = set()
    patterns = []
    for row in rates_table:
        if not isinstance(row, dict):
            continue
        key_tuple = tuple(str(row.get(k, "") or "").strip() for k in stable_keys)
        if key_tuple in seen:
            continue
        seen.add(key_tuple)
        pattern = {k: str(row.get(k, "") or "").strip() for k in stable_keys}
        patterns.append(pattern)
    patterns = enrich_patterns_with_group_constant_postals(patterns, stable_keys, rates_table)
    return patterns, stable_keys, changing_keys, distinct_count


def export_rate_card_full_json(path: Path) -> dict | None:
    """
    Same JSON structure as main() writes to ``output/{stem}_rate_card.json`` (no file write).
    Returns None if there is no Rate card sheet.
    """
    wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
    ws_general = _find_sheet(wb, GENERAL_INFO_SHEET)
    ws_rate_card = _find_sheet(wb, RATE_CARD_SHEET)
    if ws_rate_card is None:
        return None
    general_info = extract_general_info(ws_general)
    conditional_rules = extract_conditional_rules(ws_rate_card)
    costs, rates_table = extract_costs_and_rates_table(ws_rate_card)
    patterns, pattern_keys_stable, pattern_keys_changing, distinct_count_by_key = extract_patterns(rates_table)
    return {
        "general_info": general_info,
        "conditional_rules": conditional_rules,
        "costs": costs,
        "rates_table": rates_table,
        "pattern_analysis": {
            "stable_keys": pattern_keys_stable,
            "changing_keys": pattern_keys_changing,
            "distinct_count_by_key": distinct_count_by_key,
        },
        "patterns": patterns,
    }


def main():
    if not INPUT_DIR.exists():
        print(f"Folder not found: {INPUT_DIR}")
        return
    files = sorted(INPUT_DIR.glob("*.xlsx"))
    if not files:
        print("No .xlsx files in 'previous rate' folder.")
        return
    print("Files in 'previous rate' folder:")
    for i, f in enumerate(files, 1):
        print(f"  {i}. {f.name}")
    try:
        idx = int(input("Which file number? ").strip())
        path = files[idx - 1]
    except (ValueError, IndexError):
        print("Invalid choice.")
        return

    out = export_rate_card_full_json(path)
    if not out:
        print("No 'Rate card' sheet found in the workbook.")
        return

    general_info = out["general_info"]
    conditional_rules = out["conditional_rules"]
    costs = out["costs"]
    rates_table = out["rates_table"]
    pattern_keys_stable = out["pattern_analysis"]["stable_keys"]
    pattern_keys_changing = out["pattern_analysis"]["changing_keys"]
    patterns = out["patterns"]

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = OUTPUT_DIR / f"{path.stem}_rate_card.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(out, f, indent=2, ensure_ascii=False)

    print(f"\nSaved to: {out_path}")
    print(f"  General: rate_id={general_info.get('rate_id')}, validity_period={general_info.get('validity_period')}")
    print(f"  Conditional rules: {len(conditional_rules)}")
    print(f"  Costs: {len(costs)}")
    print(f"  Rates table rows: {len(rates_table)}")
    print(f"  Pattern keys (stable): {pattern_keys_stable}")
    print(f"  Pattern keys (changing): {pattern_keys_changing}")
    print(f"  Patterns: {len(patterns)}")


if __name__ == "__main__":
    main()
