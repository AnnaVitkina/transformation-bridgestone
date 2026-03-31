"""
Build a rate matrix (DataFrame / Excel) from:
  1) Tariff sheet: same pipeline as **rate_to_json.py** (open sheet → sheet_to_json → _clean_data).
     Not a separate run of clean_rate_json.py — identical cleaning. Input: `input/` folder xlsx.
  2) Example rate card from **previous rate/** (same as rate_card_extraction.py → patterns).

Origin / return workflow (and optional **flat-zero** when the rate card has **3** patterns):
  - **1–2 patterns:** choose origin vs return (or origin only if a single pattern).
  - **3 patterns:** choose **origin**, **flat-zero**, and **return** (each 1–3, all different). Then choose
    **zero-base**: copy **p/unit** (and per-weight) amounts from the **origin** or **return** tariff table;
    **Flat**-measure bands (and FTL / synthetic ``>=``) are filled with **0**. Rows are emitted in order:
    origin → flat-zero → return.
  - **Origin table**: first tariff block whose title does NOT contain "return" (fallback: first block).
  - **Return table**: first block whose title contains "return" (fallback: second block, or first if only one).
  - Origin / return pattern fields are filled from the corresponding tariff table (Zip 2 / column labels).

Pattern columns come first. Empty fields are filled from tariff headers when labels match, from
Zip 2 → city rules, from **lane columns** (non–weight-band columns left of `<10,99`-style bands,
e.g. “Area Code”) for rate-card **changing** keys, and from the rate card table when a value is
constant for all rows of that pattern (postal codes that mix outbound/return in one sheet).

Debug: set environment variable ``RATE_CREATION_DEBUG=1`` (or ``yes``/``true``) for optional verbose
trace (tables, patterns, column mapping). No interactive debug prompt at startup.

Synthetic ``>=X`` FLAT columns: you are prompted at matrix build time (default **yes**). Set
``RATE_ADD_SYNTHETIC_GEQ=0`` or ``no`` to skip without prompting; ``1``/``yes`` forces add.
If the tariff has an **FTL** column and you add synthetics, a second prompt asks whether to fill
those cells by **calculation** (weight-band math) or by copying the **FTL** column per row.
``RATE_SYNTHETIC_GEQ_FILL=calculation`` or ``ftl`` skips that prompt.

Each run also writes JSON snapshots under ``processing/``: tariff rows (same as ``rate_to_json.py``
after keyword table selection), ``*_rate_card.json`` (same as ``rate_card_extraction.py``), and a
copy of the final ``*_rate_matrix.json``.

When the rate card lists multiple **Transport cost** blocks, shipment/lane rows appear **once**. In
Excel, each cost is a **horizontal block**: rows 1–4 hold name / applies if / rate by / calculation
y
rule (merged over that block), then **Currency + weight bands** (+ optional **FTL** column) repeat per
cost. **Currency** is parsed from band cells (e.g. ``6.769 €`` → EUR, amounts without symbols).
Weight bracket **headers**: ``<10,99`` / ``0-10`` → ``<11`` (when not paired with ``>``); **``a-b``**
with **``> b``** in the same block (e.g. ``8251-10000`` + ``> 10000``) → **``<b>``** only, no
**``>=b+1``** synthetic. **``>X``** (p/unit) → **``<X.001``**; synthetic **``>=(X+0.001)``** FLAT
(rate × X ÷ per-unit divisor; **``p/100 unit``** → ÷100). Otherwise last **``<``** non-Flat band gets **``>=N``**
(rate × N ÷ divisor). Shipment/
pattern **titles** sit on the **same row** as Flat / p/unit. Costs named like **Transport cost (FTL)**
are ordered after non-FTL costs and use only **Currency + FTL** when the tariff has an FTL column;
other costs omit the FTL column.
JSON stores `transport_costs`, `rate_block_columns`, `rate_block_columns_by_cost`, optional
`rate_column_measures`, and a wide `matrix` with keys ``__bk{i}_c{j}__``.
"""
from __future__ import annotations

import json
import math
import os
import re
import warnings
from pathlib import Path

# Verbose trace only when RATE_CREATION_DEBUG=1 (or true/yes/y). No stdin prompt.
_DEBUG = False


def _init_debug_from_env() -> None:
    """Set ``_DEBUG`` from ``RATE_CREATION_DEBUG`` only (no interactive question)."""
    global _DEBUG
    env = os.environ.get("RATE_CREATION_DEBUG", "").strip().lower()
    _DEBUG = env in ("1", "true", "yes", "y")
    if _DEBUG:
        print("[rate_creation] Verbose debug is ON (RATE_CREATION_DEBUG).\n")


def _ui_heading(title: str) -> None:
    """Clear section separator for interactive flow."""
    print("\n" + "=" * 64)
    print(f"  {title}")
    print("=" * 64)


def _prompt_add_synthetic_geq_columns() -> bool:
    """
    Ask whether to add code-generated FLAT columns whose headers look like '>=X' (after '>X'
    and after the last '<' weight band). Env ``RATE_ADD_SYNTHETIC_GEQ`` = 1/0 or yes/no skips the prompt.
    """
    env = os.environ.get("RATE_ADD_SYNTHETIC_GEQ", "").strip().lower()
    if env in ("1", "true", "yes", "y"):
        print("[rate_creation] RATE_ADD_SYNTHETIC_GEQ=yes — extra '>=' FLAT columns will be added.\n")
        return True
    if env in ("0", "false", "no", "n"):
        print("[rate_creation] RATE_ADD_SYNTHETIC_GEQ=no — skipping extra '>=' FLAT columns.\n")
        return False
    try:
        q = input("Add extra manually calculated FLAT columns [Y/n]: ").strip().lower()
    except EOFError:
        print("  (EOF — default: yes)\n")
        return True
    if not q:
        return True
    return q not in ("n", "no", "0")


def _prompt_synthetic_geq_fill_mode() -> str:
    """
    When an FTL tariff column exists: how to fill synthetic '>=' FLAT cells.
    Returns ``'calculation'`` (rate × weight / divisor) or ``'ftl'`` (copy row value from FTL column).

    Env ``RATE_SYNTHETIC_GEQ_FILL`` = ``calculation`` | ``ftl`` skips the prompt.
    """
    env = os.environ.get("RATE_SYNTHETIC_GEQ_FILL", "").strip().lower()
    if env in ("ftl", "f"):
        print("[rate_creation] RATE_SYNTHETIC_GEQ_FILL=ftl — '>=' cells copy the FTL column per row.\n")
        return "ftl"
    if env in ("calculation", "calc", "c", "1"):
        print("[rate_creation] RATE_SYNTHETIC_GEQ_FILL=calculation — '>=' cells from weight-band math.\n")
        return "calculation"
    try:
        print(
            "This tariff has an FTL column. How should the new synthetic '>=' FLAT cells be filled?\n"
            "  [1] Calculation — multiply rates by weight limits (same rules as other bands)\n"
            "  [2] FTL         — copy the FTL cell value on each row\n"
        )
        q = input("Choose 1 or 2 (default 1): ").strip().lower()
    except EOFError:
        print("  (EOF — default: calculation)\n")
        return "calculation"
    if q in ("2", "ftl", "f"):
        return "ftl"
    return "calculation"


def _debug_line(msg: str, indent: int = 0) -> None:
    if not _DEBUG:
        return
    pad = "  " * (indent + 1)
    print(f"{pad}{msg}")


def _debug_step(step: int, title: str) -> None:
    if not _DEBUG:
        return
    print(f"\n{'=' * 72}")
    print(f"  DEBUG STEP {step}: {title}")
    print(f"{'=' * 72}")


try:
    import openpyxl
    import pandas as pd
except ImportError:
    print("Please install: pip install openpyxl pandas")
    raise SystemExit(1)

from rate_to_json import (
    INPUT_DIR as TARIFF_INPUT_DIR,
    _clean_data,
    _extract_tables_by_name,
    _get_table_title_from_row,
    _split_into_tables,
    clean_tariff_rows_twice,
    combine_tariff_sheet_rows,
    combined_tariff_output_basename,
    export_tariff_sheet_rows,
    get_xlsx_files,
    parse_tariff_file_index_list,
    sheet_to_json,
)
from rate_card_extraction import INPUT_DIR as PREVIOUS_RATE_DIR, export_rate_card_full_json

OUTPUT_DIR = Path(__file__).resolve().parent / "output"
# Intermediate JSON snapshots (same payloads as rate_to_json / rate_card_extraction / final matrix)
PROCESSING_DIR = Path(__file__).resolve().parent / "processing"


def _cell_is_empty_for_column_drop(v) -> bool:
    """True if value counts as empty for 'all-empty column' removal."""
    if v is None:
        return True
    try:
        if pd.isna(v):
            return True
    except (TypeError, ValueError):
        pass
    if isinstance(v, str):
        s = v.strip()
        return not s or s.lower() in ("nan", "none")
    s = str(v).strip()
    return not s or s.lower() in ("nan", "none", "<na>")


def drop_all_empty_columns(
    df: pd.DataFrame,
    *,
    keep_columns: frozenset[str] | None = None,
) -> pd.DataFrame:
    """Remove columns where every cell is empty (NaN, None, blank, or string 'nan'/'none')."""
    if df.empty or len(df.columns) == 0:
        return df
    keep_set = keep_columns or frozenset()
    keep = [
        c
        for c in df.columns
        if c in keep_set or not df[c].map(_cell_is_empty_for_column_drop).all()
    ]
    return df[keep]


def _numeric_key_from_bracket_header(label: str) -> float:
    """Sort key for weight-band column titles (e.g. '<10,99', '0-10', '> 10000')."""
    base = _label_base_for_match(label or "").strip()
    t = base.lower().replace(",", ".")
    m = re.search(r"(\d+(?:\.\d+)?)", t)
    if m:
        return float(m.group(1))
    return 0.0


def sort_extra_tariff_columns(cols: list[str]) -> list[str]:
    """
    Keep non–weight-band columns in order (lane ids, etc.), then sort band columns by numeric threshold.
    Matches typical Advanced Export tariff table column order.
    """
    non_band: list[str] = []
    band: list[str] = []
    for c in cols:
        if _is_weight_band_label(c):
            band.append(c)
        else:
            non_band.append(c)
    band_sorted = sorted(band, key=lambda x: _numeric_key_from_bracket_header(x))
    return non_band + band_sorted


def _cost_applies_if_as_string(cost: dict) -> str:
    applies = cost.get("applies_if")
    if isinstance(applies, list):
        return "\n".join(str(x) for x in applies if x is not None and str(x).strip())
    if applies is not None:
        return str(applies).strip()
    return ""


def _is_weight_band_label(label: str) -> bool:
    """True if header cell is a weight/price band column (e.g. '<10,99', '0-10', '>1000'), not a lane id."""
    base = _label_base_for_match(label or "").strip()
    if not base:
        return False
    if _cell_is_column_role_label(base):
        return False
    t = base.lower()
    if re.match(r"^<\d", t):
        return True
    if re.match(r"^\d+\s*[-–]\s*\d+", t):
        return True
    if re.match(r"^>\s*\d", t.replace(" ", "")):
        return True
    if re.search(r"\d+\s*[-–]\s*\d+\s*kg", t):
        return True
    return False


def _is_geq_synthetic_bracket_label(label: str) -> bool:
    """Synthetic column e.g. '>=12500' added after last p/unit band."""
    return str(_label_base_for_match(label or "").strip()).startswith(">=")


def _is_flat_tariff_measure(measure: str | None) -> bool:
    """True only for shipment-style Flat; per-weight columns use p/unit, p/100 unit, p/10 unit, …"""
    return str(measure or "").strip().lower() == "flat"


def _label_is_flat_zero_tariff_band(
    label: str,
    measure_by_label: dict[str, str],
) -> bool:
    """
    For **flat-zero** matrix rows: these columns get **0**; **p/unit** (and p/X unit) copy from the
    zero-base tariff row. Currency passes through; FTL and synthetic ``>=`` are treated like Flat → 0.
    """
    if _is_currency_column_name(label):
        return False
    if _is_ftl_tariff_column(label):
        return True
    if _is_geq_synthetic_bracket_label(label):
        return True
    base = _label_base_for_match(label)
    m = measure_by_label.get(label) or measure_by_label.get(base) or ""
    return _is_flat_tariff_measure(m)


def _per_weight_divisor_from_measure(measure: str | None) -> float:
    """
    For per-weight tariffs, total cost for weight W kg is rate * W / divisor.
    ``p/unit`` → divisor 1. ``p/100 unit`` → 100 (rate is per 100 kg). ``p/10 unit`` → 10.
    Unknown / non-matching measures default to 1.
    """
    s = str(measure or "").strip().lower()
    if s == "p/unit" or not s:
        return 1.0
    m = re.match(r"^p/([\d.,]+)\s*unit$", s)
    if m:
        try:
            d = float(m.group(1).replace(",", "."))
            return d if d > 0 else 1.0
        except ValueError:
            return 1.0
    return 1.0


def _parse_bracket_gt_lower_bound(label: str) -> float | None:
    """
    Lower bound X from '> 10000', '>10000', '>10,99' (EU decimal comma).
    Used to show '<X.001' and add synthetic '>=(X+0.001)' FLAT with value = rate * X.
    """
    base = _label_base_for_match(label or "").strip()
    if not base:
        return None
    t = base.replace(" ", "")
    if not t.startswith(">"):
        return None
    inner = t[1:]
    if inner.count(",") == 1 and inner.rfind(",") > max(inner.rfind("."), -1):
        parts = inner.split(",")
        if len(parts) == 2 and len(parts[1]) <= 2 and parts[1].isdigit():
            whole, frac = parts[0], parts[1]
            inner = whole.replace(".", "") + "." + frac
        else:
            inner = inner.replace(",", ".")
    else:
        inner = inner.replace(",", ".")
    try:
        return float(inner)
    except ValueError:
        return None


def _float_str_trim_trailing_zeros(v: float) -> str:
    s = f"{v:.12f}".rstrip("0").rstrip(".")
    return s if s else "0"


def _format_bracket_gt_as_lt_dot_001(lower: float) -> str:
    """'> 10000' p/unit → display '<10000.001'."""
    return f"<{_float_str_trim_trailing_zeros(lower + 0.001)}"


def _synthetic_geq_name_from_gt_lower(lower: float) -> str:
    """>=10000.001' style column name."""
    return f">={_float_str_trim_trailing_zeros(lower + 0.001)}"


def _parse_bracket_upper_threshold(label: str) -> float | None:
    """
    Upper weight bound for ``<10,99`` / ``<12499,99`` (EU decimal comma) and hyphen ranges
    ``8251-10000``, ``7201-8000,99`` — hyphen parsing is delegated to :func:`_hyphen_range_upper_float`.
    Used for display ``<N`` and synthetic ``>=N`` FLAT column.
    """
    base = _label_base_for_match(label or "").strip()
    if not base:
        return None
    t = base.replace(" ", "")
    if t.startswith("<"):
        inner = t[1:]
        if inner.count(",") == 1 and inner.rfind(",") > max(inner.rfind("."), -1):
            parts = inner.split(",")
            if len(parts) == 2 and len(parts[1]) <= 2 and parts[1].isdigit():
                whole, frac = parts[0], parts[1]
                inner = whole.replace(".", "") + "." + frac
            else:
                inner = inner.replace(",", ".")
        else:
            inner = inner.replace(",", ".")
        try:
            return float(inner)
        except ValueError:
            return None
    hu = _hyphen_range_upper_float(base)
    if hu is not None:
        return hu
    return None


def _less_than_display_int_from_upper(ub: float) -> int:
    """
    '<10,99' / 0-10 → '<11'; '<12499,99' → '<12500'.
    Next integer above the open upper bound (10 → 11; 12499.99 → 12500).
    """
    if ub == int(ub):
        return int(ub) + 1
    return int(math.ceil(ub))


def _synthetic_geq_int_from_last_lt_upper(ub: float) -> int:
    """
    Integer **N** for the synthetic ``>=N`` FLAT column after the last ``<`` p/unit band (and for
    ``rate × N / divisor``). **Does not** replace :func:`_less_than_display_int_from_upper`, which
    stays used only for ``<…`` display labels.

    EU tariffs often end a band at ``12500,99`` (parsed ``12500.99``). Using :func:`math.ceil` would
    yield ``>=12501``, while a manually added bracket is typically ``>=12500``. When the integer part
    is a **round hundred** and the fractional part is **~0.99**, use ``N = int(ub)`` so the
    synthetic column matches that manual ``>=12500`` naming. Other bounds (e.g. ``12499,99`` →
    ``ceil`` → ``12500``) keep the previous behaviour.
    """
    iu = int(ub)
    frac = ub - iu
    if (
        ub != float(iu)
        and iu % 100 == 0
        and frac >= 0.99 - 1e-6
    ):
        return iu
    return _less_than_display_int_from_upper(ub)


def _hyphen_range_upper_float(label: str) -> float | None:
    """
    Upper bound (second number) of ``a-b`` weight ranges. ``(ColumnN)`` is stripped by
    :func:`_label_base_for_match` first.

    One pattern allows optional ``.,`` decimals on **both** sides; the **upper** used everywhere
    is the second bound only:

    - ``8251-10000`` → 10000
    - ``7201-8000,99`` / ``7201-8000.99`` → 8000.99
    - ``10401-12500,99`` → 12500.99

    Used by :func:`_parse_bracket_upper_threshold` (hyphen branch) and paired-``>`` display logic.
    """
    base = _label_base_for_match(label or "").strip()
    t = base.replace(" ", "")
    m = re.match(
        r"^(\d+)(?:[.,](\d+))?\s*[-–]\s*(\d+)(?:[.,](\d+))?$",
        t,
    )
    if not m:
        return None
    _a1, _f1, a2, f2 = m.group(1), m.group(2), m.group(3), m.group(4)
    if f2 is None:
        return float(a2)
    return float(f"{a2}.{f2}")


def _gt_lower_matches_hyphen_upper(gl: float, upper: float) -> bool:
    """
    ``> 10000`` pairs ``8251-10000`` (exact). ``> 8000`` pairs ``7201-8000,99`` (GT at integer
    boundary, upper 8000.99).
    """
    if abs(gl - upper) < 1e-6:
        return True
    if abs(upper - int(upper)) < 1e-9:
        return abs(gl - upper) < 1e-6
    # Fractional upper (e.g. 8000.99): match ``> 8000`` with gl == 8000 == int(upper)
    return abs(gl - int(upper)) < 1e-9


def _hyphen_paired_lt_display(upper: float) -> str:
    """
    Paired with a ``> …`` column in the same block: integer upper (e.g. 10000) → ``<10000``;
    fractional upper (e.g. 8000.99) → ``<8001`` via :func:`_less_than_display_int_from_upper`.
    """
    if abs(upper - int(upper)) < 1e-9:
        return f"<{int(upper)}"
    return f"<{_less_than_display_int_from_upper(upper)}"


def _hyphen_range_lt_display_when_paired_with_gt(
    label: str, rate_block_context: list[str]
) -> str | None:
    """
    When 'a-b' (e.g. 8251-10000) appears with '> b' in the same block, show '<b>' (e.g. <10000),
    not '<b+1>'. Suppresses the extra '>=10001' synthetic via Phase B skip.
    EU ranges like 7201-8000,99 with '> 8000' → <8001.
    """
    upper = _hyphen_range_upper_float(label)
    if upper is None:
        return None
    for c in rate_block_context:
        if c == label:
            continue
        gl = _parse_bracket_gt_lower_bound(c)
        if gl is not None and _gt_lower_matches_hyphen_upper(gl, upper):
            return _hyphen_paired_lt_display(upper)
    return None


def _hyphen_range_paired_with_matching_gt_column(src: str, rbc: list[str]) -> bool:
    """True if src is e.g. '8251-10000' or '7201-8000,99' and some other column is a matching '> …'."""
    upper = _hyphen_range_upper_float(src)
    if upper is None:
        return False
    for c in rbc:
        if c == src:
            continue
        gl = _parse_bracket_gt_lower_bound(c)
        if gl is not None and _gt_lower_matches_hyphen_upper(gl, upper):
            return True
    return False


def format_bracket_header_display(
    label: str, *, rate_block_context: list[str] | None = None
) -> str:
    """
    '<10,99' / '0-10' → '<11'; '> 10000' p/unit → '<10000.001' (see synthetic >= column).
    '8251-10000' + '> 10000' in the same block → '<10000' (not '<10001'); no '>=' synthetic for that band.
    Currency, FTL, and synthetic '>=…' pass through unchanged.
    """
    if _is_currency_column_name(label) or _is_ftl_tariff_column(label):
        return label
    if _is_geq_synthetic_bracket_label(label):
        return label
    if not _is_weight_band_label(label):
        return label
    if rate_block_context is not None:
        paired_lt = _hyphen_range_lt_display_when_paired_with_gt(label, rate_block_context)
        if paired_lt is not None:
            return paired_lt
    gt_lb = _parse_bracket_gt_lower_bound(label)
    if gt_lb is not None:
        return _format_bracket_gt_as_lt_dot_001(gt_lb)
    ub = _parse_bracket_upper_threshold(label)
    if ub is None:
        return label
    n = _less_than_display_int_from_upper(ub)
    return f"<{n}"


def _is_ftl_tariff_column(label: str) -> bool:
    """Tariff column whose header is FTL (full truckload), not a weight bracket."""
    return _label_base_for_match(label or "").strip().upper() == "FTL"


def build_rate_block_column_order(extra_cols: list[str]) -> tuple[list[str], list[str]]:
    """
    Split tariff extra columns into (lane_misc, rate_block).
    Rate block = Currency + weight bands (sorted) + FTL column(s) if present; lane_misc = rest.
    """
    curr = [c for c in extra_cols if str(c).strip().lower() == "currency"]
    ftl_cols = [c for c in extra_cols if c not in curr and _is_ftl_tariff_column(c)]
    bands = [c for c in extra_cols if c not in curr and _is_weight_band_label(c)]
    bands_sorted = sorted(bands, key=lambda x: _numeric_key_from_bracket_header(x))
    rate_block = curr + bands_sorted + ftl_cols
    lane_misc = [c for c in extra_cols if c not in rate_block]
    return lane_misc, rate_block


def _cost_name_implies_ftl_transport(cost: dict) -> bool:
    """True if rate card cost name references FTL, e.g. 'Transport cost (FTL)'."""
    return bool(re.search(r"\bFTL\b", str(cost.get("name") or ""), re.I))


def sort_transport_costs_non_ftl_first(costs: list[dict]) -> list[dict]:
    """Non–FTL-named costs first, then FTL-named (stable within each group)."""
    non = [c for c in costs if isinstance(c, dict) and not _cost_name_implies_ftl_transport(c)]
    ftl = [c for c in costs if isinstance(c, dict) and _cost_name_implies_ftl_transport(c)]
    return non + ftl


def rate_block_cols_for_transport_cost(cost: dict, rate_block_cols: list[str]) -> list[str]:
    """
    LTL / generic costs: all rate columns except the FTL tariff column.
    FTL-named costs: Currency + FTL column only (when tariff has an FTL column); else full block.
    """
    has_ftl_col = any(_is_ftl_tariff_column(c) for c in rate_block_cols)
    if _cost_name_implies_ftl_transport(cost):
        if has_ftl_col:
            out = [c for c in rate_block_cols if _is_currency_column_name(c) or _is_ftl_tariff_column(c)]
            return out if out else list(rate_block_cols)
        return list(rate_block_cols)
    out = [c for c in rate_block_cols if not _is_ftl_tariff_column(c)]
    return out if out else list(rate_block_cols)


def _is_currency_column_name(name: str) -> bool:
    return str(name or "").strip().lower() == "currency"


def _rate_band_columns(rate_block_cols: list[str]) -> list[str]:
    """Columns that hold numeric rates (exclude Currency)."""
    return [c for c in rate_block_cols if not _is_currency_column_name(c)]


# ISO-like codes and symbols → normalized code for the Currency column
_CURRENCY_CODE_WORDS: tuple[str, ...] = (
    "EUR",
    "USD",
    "GBP",
    "PLN",
    "CHF",
    "CZK",
    "HUF",
    "SEK",
    "NOK",
    "DKK",
    "RON",
    "BGN",
    "TRY",
    "CNY",
    "JPY",
    "INR",
    "AUD",
    "CAD",
    "NZD",
    "MXN",
    "BRL",
    "ZAR",
    "AED",
    "SAR",
    "ILS",
    "KRW",
    "THB",
    "SGD",
    "MYR",
    "IDR",
    "PHP",
    "VND",
    "TWD",
    "HKD",
)


def _parse_currency_and_strip_amount(val) -> tuple[str | None, str | float | None]:
    """
    From a tariff cell like '6.769 €', '64.02 PLN', '12.50 USD', return (code, amount_without_unit).
    If no currency token is found, returns (None, original value unchanged).
    """
    if val is None:
        return None, None
    try:
        if pd.isna(val):
            return None, None
    except (TypeError, ValueError):
        pass
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        return None, val
    s0 = str(val).strip()
    if not s0:
        return None, val

    s = s0
    found_code: str | None = None

    # Explicit 3-letter codes (word boundary)
    for code in _CURRENCY_CODE_WORDS:
        m = re.search(r"(?<!\w)" + re.escape(code) + r"(?!\w)", s, re.I)
        if m:
            found_code = code.upper()
            s = (s[: m.start()] + s[m.end() :]).strip()
            break

    if found_code is None:
        low = s.lower()
        # Czech koruna in tariffs (Excel often uses localized suffix, not ISO "CZK")
        if "kč" in low:
            found_code = "CZK"
            s = re.sub(r"kč", "", s, flags=re.I).strip()
        elif "zł" in low or "zl" in low.replace("ł", "l"):
            found_code = "PLN"
            s = re.sub(r"zł|zl", "", s, flags=re.I).strip()
        elif "€" in s:
            found_code = "EUR"
            s = s.replace("€", "").strip()
        elif re.search(r"(?<!\w)£(?!\w)", s):
            found_code = "GBP"
            s = re.sub(r"£", "", s).strip()
        elif "$" in s:
            found_code = "USD"
            s = s.replace("$", "").strip()

    if found_code is None:
        return None, val

    # Normalize amount: spaces, NBSP; keep digits, one decimal separator
    s = s.replace("\u00a0", " ").replace(" ", "")
    s = re.sub(r"[^\d,.\-]", "", s)
    if not s or s in ("-", "."):
        return found_code, None
    # Prefer last comma or dot as decimal if both present
    if s.count(",") and s.count("."):
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif s.count(",") == 1 and s.count(".") == 0:
        parts = s.split(",")
        if len(parts[-1]) <= 2 and parts[0].replace("-", "").isdigit():
            s = parts[0] + "." + parts[1]
        else:
            s = s.replace(",", "")
    elif s.count(",") > 1:
        s = s.replace(",", "")
    try:
        return found_code, float(s)
    except ValueError:
        return found_code, s if s else None


def normalize_currency_column_and_strip_band_amounts(
    df: pd.DataFrame,
    rate_block_cols: list[str],
) -> tuple[pd.DataFrame, list[str]]:
    """
    Ensure a **Currency** column exists as the first rate column; fill it by parsing band cells
    (e.g. '6.769 €' → EUR + 6.769) and strip currency symbols from band values.
    """
    if df.empty or not rate_block_cols:
        return df, rate_block_cols

    df = df.copy()
    curr_name = next((c for c in df.columns if _is_currency_column_name(str(c))), None)
    rbc = list(rate_block_cols)

    if curr_name is None:
        curr_name = "Currency"
        df[curr_name] = None
        if not rbc or not _is_currency_column_name(rbc[0]):
            rbc.insert(0, curr_name)
    else:
        if rbc and rbc[0] != curr_name:
            rbc = [c for c in rbc if c != curr_name]
            rbc.insert(0, curr_name)

    bands = _rate_band_columns(rbc)
    for bc in bands:
        if bc in df.columns:
            df[bc] = df[bc].astype(object)
    if curr_name in df.columns:
        df[curr_name] = df[curr_name].astype(object)

    for idx in df.index:
        row_currency: str | None = None
        if curr_name in df.columns:
            existing = df.at[idx, curr_name]
            if existing is not None and str(existing).strip():
                row_currency = str(existing).strip().upper()
        for bc in bands:
            if bc not in df.columns:
                continue
            cell = df.at[idx, bc]
            code, stripped = _parse_currency_and_strip_amount(cell)
            if code and row_currency is None:
                row_currency = code
            if stripped is not None and stripped != cell:
                df.at[idx, bc] = stripped
        if curr_name in df.columns and row_currency:
            if not (df.at[idx, curr_name] is not None and str(df.at[idx, curr_name]).strip()):
                df.at[idx, curr_name] = row_currency

    return df, rbc


def add_synthetic_geq_flat_after_last_p_unit(
    df: pd.DataFrame,
    rate_block_cols: list[str],
    rate_column_measures: list[str],
    *,
    fill_synthetic_from_ftl: bool = False,
) -> tuple[pd.DataFrame, list[str], list[str]]:
    """
    (1) Each **'>X'** column whose measure is **not** Flat (p/unit, p/100 unit, …): insert
        **'>=(X+0.001)'** FLAT with value = (rate in '>X') * X / divisor, where divisor is 1 for
        ``p/unit`` and X for ``p/X unit`` (e.g. p/100 unit → divide by 100).

    (2) If the last **'<'** weight-band column with a non-Flat measure (ignoring '>X' columns) has
        bracket '<…' or range, insert **'>=N'** FLAT with value = rate * N / divisor (e.g. p/100 unit
        and N=10000 → rate×10000/100).

    If ``fill_synthetic_from_ftl`` is True and an **FTL** column exists in the rate block, each new
    synthetic ``>=`` cell **copies that row's FTL value** instead of the calculated amount.
    """
    if df.empty or not rate_block_cols or not rate_column_measures:
        return df, rate_block_cols, rate_column_measures
    if len(rate_block_cols) != len(rate_column_measures):
        return df, rate_block_cols, rate_column_measures

    df = df.copy()
    rbc = list(rate_block_cols)
    meas = list(rate_column_measures)

    ftl_col: str | None = None
    if fill_synthetic_from_ftl:
        ftl_col = next((c for c in rbc if _is_ftl_tariff_column(c)), None)
        if ftl_col is None or ftl_col not in df.columns:
            fill_synthetic_from_ftl = False

    def _copy_ftl_or_compute(idx, compute_prod: float) -> object:
        """Return cell value for synthetic column: FTL copy or rounded ``compute_prod``."""
        if fill_synthetic_from_ftl and ftl_col is not None:
            raw = df.at[idx, ftl_col] if ftl_col in df.columns else None
            if raw is None or (isinstance(raw, float) and pd.isna(raw)):
                return None
            return raw
        if abs(compute_prod - round(compute_prod)) < 1e-6:
            return int(round(compute_prod))
        return round(compute_prod, 2)

    gt_to_add: list[tuple[int, str, float, str]] = []
    for i, col in enumerate(rbc):
        if _is_currency_column_name(col) or _is_ftl_tariff_column(col) or _is_geq_synthetic_bracket_label(col):
            continue
        if _is_flat_tariff_measure(meas[i]):
            continue
        lb = _parse_bracket_gt_lower_bound(col)
        if lb is None:
            continue
        new_name = _synthetic_geq_name_from_gt_lower(lb)
        if new_name in rbc:
            continue
        gt_to_add.append((i, col, lb, new_name))

    for _orig_i, src_col, lb, new_name in sorted(gt_to_add, key=lambda x: -x[0]):
        if new_name in rbc:
            continue
        pos = rbc.index(src_col)
        if src_col in df.columns:
            df[src_col] = df[src_col].astype(object)
        out_vals: list[object] = []
        for idx in df.index:
            if fill_synthetic_from_ftl:
                out_vals.append(_copy_ftl_or_compute(idx, 0.0))
                continue
            v = df.at[idx, src_col] if src_col in df.columns else None
            try:
                if v is None or (isinstance(v, float) and pd.isna(v)):
                    out_vals.append(None)
                    continue
                fv = float(v)
            except (TypeError, ValueError):
                out_vals.append(None)
                continue
            div = _per_weight_divisor_from_measure(meas[pos])
            prod = fv * float(lb) / div
            out_vals.append(_copy_ftl_or_compute(idx, prod))
        df[new_name] = out_vals
        rbc = rbc[: pos + 1] + [new_name] + rbc[pos + 1 :]
        meas = meas[: pos + 1] + ["Flat"] + meas[pos + 1 :]

    last_pu: int | None = None
    for i in range(len(rbc) - 1, -1, -1):
        col = rbc[i]
        if _is_currency_column_name(col) or _is_ftl_tariff_column(col) or _is_geq_synthetic_bracket_label(col):
            continue
        if not _is_weight_band_label(col):
            continue
        if _parse_bracket_gt_lower_bound(col) is not None:
            continue
        if _is_flat_tariff_measure(meas[i]):
            continue
        last_pu = i
        break
    if last_pu is None:
        return df, rbc, meas

    src = rbc[last_pu]
    ub = _parse_bracket_upper_threshold(src)
    if ub is None:
        return df, rbc, meas
    # e.g. 8251-10000 next to >10000: <10000.001 / >=10000.001 from '>' only — not >=10001 from range
    if _hyphen_range_paired_with_matching_gt_column(src, rbc):
        return df, rbc, meas
    n_mult = _synthetic_geq_int_from_last_lt_upper(ub)
    new_name = f">={n_mult}"
    if new_name in rbc:
        return df, rbc, meas

    if src in df.columns:
        df[src] = df[src].astype(object)
    out_vals2: list[object] = []
    for idx in df.index:
        if fill_synthetic_from_ftl:
            out_vals2.append(_copy_ftl_or_compute(idx, 0.0))
            continue
        v = df.at[idx, src] if src in df.columns else None
        try:
            if v is None or (isinstance(v, float) and pd.isna(v)):
                out_vals2.append(None)
                continue
            fv = float(v)
        except (TypeError, ValueError):
            out_vals2.append(None)
            continue
        div = _per_weight_divisor_from_measure(meas[last_pu])
        prod = fv * float(n_mult) / div
        out_vals2.append(_copy_ftl_or_compute(idx, prod))
    df[new_name] = out_vals2

    new_rbc = list(rbc[: last_pu + 1]) + [new_name] + list(rbc[last_pu + 1 :])
    new_meas = list(meas[: last_pu + 1]) + ["Flat"] + list(meas[last_pu + 1 :])
    return df, new_rbc, new_meas


def widen_matrix_for_horizontal_cost_blocks(
    df: pd.DataFrame,
    transport_costs: list[dict],
    shipment_cols: list[str],
    rate_block_cols: list[str],
    *,
    rate_block_cols_by_cost: list[list[str]] | None = None,
) -> tuple[pd.DataFrame, list[str]]:
    """
    Duplicate rate columns once per transport cost; keys ``__bk{i}_c{j}__``.
    Each cost may use a different subset of columns (e.g. LTL blocks omit FTL; FTL cost uses Currency + FTL).
    """
    costs = [c for c in transport_costs if isinstance(c, dict)]
    if not costs or not rate_block_cols:
        ord0 = [c for c in shipment_cols if c in df.columns] + [c for c in rate_block_cols if c in df.columns]
        return df, ord0

    n_b = len(costs)
    rb_each = (
        rate_block_cols_by_cost
        if rate_block_cols_by_cost is not None
        else [list(rate_block_cols) for _ in range(n_b)]
    )
    if len(rb_each) != n_b:
        rb_each = [list(rate_block_cols) for _ in range(n_b)]

    col_order: list[str] = []
    for s in shipment_cols:
        if s in df.columns:
            col_order.append(s)
    for i in range(n_b):
        for j in range(len(rb_each[i])):
            col_order.append(f"__bk{i}_c{j}__")

    rows: list[dict] = []
    for _, row in df.iterrows():
        d: dict = {}
        for s in shipment_cols:
            if s in df.columns:
                d[s] = row[s]
        for i in range(n_b):
            for j, bname in enumerate(rb_each[i]):
                k = f"__bk{i}_c{j}__"
                d[k] = row[bname] if bname in df.columns else None
        rows.append(d)
    return pd.DataFrame(rows), col_order


def _excel_cost_block_start_columns(n_ship: int, n_rate_per_block: list[int]) -> list[int]:
    """1-based Excel column index where each transport cost block starts."""
    out: list[int] = []
    c = n_ship + 1
    for nr in n_rate_per_block:
        out.append(c)
        c += nr
    return out


def write_rate_matrix_excel_advanced_export_layout(
    df: pd.DataFrame,
    path: Path,
    *,
    column_order: list[str],
    transport_costs: list[dict] | None,
    shipment_cols: list[str],
    rate_block_cols: list[str],
    wide_format: bool,
    rate_column_measures: list[str] | None = None,
    rate_block_cols_by_cost: list[list[str]] | None = None,
    rate_column_measures_by_cost: list[list[str]] | None = None,
) -> None:
    """
    Advanced Export layout: shipment columns once on the left; each transport cost is a **block**
    to the right with rows 1–4 = Name, Applies if, Rate by, Calculation rule (merged over that
    block’s width), then header row = Currency + weight bands (width may differ per cost), optional
    measure row. When measures exist, shipment column **titles** sit on the **same row** as Flat /
    p/unit (not merged across the band header row + measure row).
    """
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    costs = [c for c in (transport_costs or []) if isinstance(c, dict)]
    data_cols = [c for c in column_order if c in df.columns]
    n_ship = len([c for c in shipment_cols if c in df.columns])
    n_cost = len(costs)
    rb_by_cost = rate_block_cols_by_cost
    if wide_format and n_cost and rb_by_cost is None:
        rb_by_cost = [list(rate_block_cols) for _ in range(n_cost)]
    n_rate_per_block = [len(rb_by_cost[i]) for i in range(n_cost)] if rb_by_cost and n_cost else [len(rate_block_cols)]
    block_starts = _excel_cost_block_start_columns(n_ship, n_rate_per_block) if n_cost else []

    meas_by_cost = rate_column_measures_by_cost
    if meas_by_cost is None and rate_column_measures and n_cost and rb_by_cost:
        m0 = rate_column_measures or []
        meas_by_cost = [list(m0) for _ in range(n_cost)] if m0 else None
    meas_u = rate_column_measures or []
    if wide_format:
        show_band_measures = bool(
            meas_by_cost
            and rb_by_cost
            and n_cost
            and all(len(meas_by_cost[i]) == len(rb_by_cost[i]) for i in range(n_cost))
            and any(any(str(x).strip() for x in meas_by_cost[i]) for i in range(n_cost))
        )
    else:
        show_band_measures = bool(
            meas_u
            and rate_block_cols
            and len(meas_u) == len(rate_block_cols)
            and any(str(x).strip() for x in meas_u)
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Matrix"
    header_fill = PatternFill("solid", fgColor="E7E6E6")
    label_fill = PatternFill("solid", fgColor="EEF0F0")
    title_fill = PatternFill("solid", fgColor="D9E1F2")

    row_cursor = 1

    if wide_format and n_cost and rb_by_cost and n_rate_per_block and all(x > 0 for x in n_rate_per_block):
        # Shipment / pattern columns: merge rows 1–4 so they align with each cost block’s 4 metadata rows
        if n_ship:
            for j in range(1, n_ship + 1):
                ws.merge_cells(start_row=row_cursor, end_row=row_cursor + 3, start_column=j, end_column=j)
                ws.cell(row=row_cursor, column=j).fill = header_fill
        # Rows 1–4: one merged band per cost block (to the right of shipment columns)
        for ri in range(4):
            rr = row_cursor + ri
            for bi, cst in enumerate(costs):
                sc = block_starts[bi]
                ec = sc + n_rate_per_block[bi] - 1
                if ec < sc:
                    continue
                ws.merge_cells(start_row=rr, end_row=rr, start_column=sc, end_column=ec)
                c = ws.cell(row=rr, column=sc)
                c.alignment = Alignment(wrap_text=True, vertical="top", horizontal="center")
                if ri == 0:
                    c.value = str(cst.get("name") or "").strip()
                    c.font = Font(bold=True, size=11)
                    c.fill = title_fill
                elif ri == 1:
                    c.value = "Applies if: " + _cost_applies_if_as_string(cst)
                    c.fill = label_fill
                elif ri == 2:
                    c.value = "Rate by: " + str(cst.get("rate_by") or "").strip()
                    c.fill = label_fill
                else:
                    c.value = "Calculation rule: " + str(cst.get("Calculation rule") or "").strip()
                    c.fill = label_fill
            ws.row_dimensions[rr].height = 32 if ri != 1 else min(120, 60 + 8 * n_cost)

        header_row = row_cursor + 4
        ship_list = [c for c in shipment_cols if c in df.columns]
        if show_band_measures:
            for j in range(1, n_ship + 1):
                c = ws.cell(row=header_row, column=j)
                c.value = None
                c.fill = header_fill
            for bi in range(n_cost):
                sc = block_starts[bi]
                rnames = rb_by_cost[bi]
                for rj, rname in enumerate(rnames):
                    col_idx = sc + rj
                    h = ws.cell(
                        row=header_row,
                        column=col_idx,
                        value=format_bracket_header_display(
                            rname, rate_block_context=rb_by_cost[bi]
                        ),
                    )
                    h.font = Font(bold=True)
                    h.fill = header_fill
            measure_row = header_row + 1
            for j, name in enumerate(ship_list, start=1):
                h = ws.cell(row=measure_row, column=j, value=name)
                h.font = Font(bold=True)
                h.fill = header_fill
                h.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for bi in range(n_cost):
                sc = block_starts[bi]
                mrow = meas_by_cost[bi]
                for rj, mtext in enumerate(mrow):
                    col_idx = sc + rj
                    c = ws.cell(row=measure_row, column=col_idx, value=mtext or None)
                    c.font = Font(italic=True, size=10)
                    c.fill = header_fill
                    c.alignment = Alignment(horizontal="center", vertical="center")
            value_row = header_row + 2
        else:
            for j, name in enumerate(ship_list, start=1):
                h = ws.cell(row=header_row, column=j, value=name)
                h.font = Font(bold=True)
                h.fill = header_fill
            for bi in range(n_cost):
                sc = block_starts[bi]
                for rj, rname in enumerate(rb_by_cost[bi]):
                    col_idx = sc + rj
                    h = ws.cell(
                        row=header_row,
                        column=col_idx,
                        value=format_bracket_header_display(
                            rname, rate_block_context=rb_by_cost[bi]
                        ),
                    )
                    h.font = Font(bold=True)
                    h.fill = header_fill
            value_row = header_row + 1
        for _, ser in df.iterrows():
            for j, colname in enumerate(data_cols, start=1):
                v = ser.get(colname)
                try:
                    if pd.isna(v):
                        v = None
                except (TypeError, ValueError):
                    pass
                ws.cell(row=value_row, column=j, value=v)
            value_row += 1
    else:
        ship_set = {c for c in shipment_cols if c in df.columns}
        if show_band_measures:
            for j, colname in enumerate(data_cols, start=1):
                if colname in ship_set:
                    c = ws.cell(row=row_cursor, column=j, value=None)
                    c.fill = header_fill
                else:
                    disp = (
                        format_bracket_header_display(
                            colname, rate_block_context=rate_block_cols
                        )
                        if colname in rate_block_cols
                        else colname
                    )
                    cell = ws.cell(row=row_cursor, column=j, value=disp)
                    cell.font = Font(bold=True)
                    cell.fill = header_fill
            mr = row_cursor + 1
            for j, colname in enumerate(data_cols, start=1):
                if colname in ship_set:
                    cell = ws.cell(row=mr, column=j, value=colname)
                    cell.font = Font(bold=True)
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                elif colname in rate_block_cols:
                    idx = rate_block_cols.index(colname)
                    val = meas_u[idx] if idx < len(meas_u) else None
                    c = ws.cell(row=mr, column=j, value=val)
                    c.font = Font(italic=True, size=10)
                    c.fill = header_fill
                else:
                    c = ws.cell(row=mr, column=j, value=None)
                    c.fill = header_fill
            r = row_cursor + 2
        else:
            for j, colname in enumerate(data_cols, start=1):
                disp = (
                    format_bracket_header_display(
                        colname, rate_block_context=rate_block_cols
                    )
                    if colname in rate_block_cols
                    else colname
                )
                cell = ws.cell(row=row_cursor, column=j, value=disp)
                cell.font = Font(bold=True)
                cell.fill = header_fill
            r = row_cursor + 1
        for _, ser in df.iterrows():
            for j, colname in enumerate(data_cols, start=1):
                v = ser.get(colname)
                try:
                    if pd.isna(v):
                        v = None
                except (TypeError, ValueError):
                    pass
                ws.cell(row=r, column=j, value=v)
            r += 1

    max_col = max(len(data_cols), 1)
    for j in range(1, max_col + 1):
        letter = get_column_letter(j)
        if j <= n_ship:
            w = 13
        elif wide_format and n_cost and block_starts and rb_by_cost:
            w = 10
            for bi, sc in enumerate(block_starts):
                if j == sc and rb_by_cost[bi] and _is_currency_column_name(rb_by_cost[bi][0]):
                    w = 9
                    break
        else:
            w = 10
        ws.column_dimensions[letter].width = w

    wb.save(path)


def _tariff_lane_column_keys(col_labels: dict[str, str]) -> list[str]:
    """
    Lane / area columns are to the **left** of the first weight-band header (e.g. Area Code before '<10,99').
    Excludes Zip 2 (handled separately for city fill).
    """
    lane: list[str] = []
    for ck in _sorted_column_keys(col_labels):
        lab = col_labels.get(ck, "") or ""
        if _label_is_zip2_lane(_label_base_for_match(lab)):
            continue
        if _is_weight_band_label(lab):
            break
        if str(lab).strip():
            lane.append(ck)
    return lane


def _rate_row_matches_stable_pattern(row: dict, pattern: dict[str, str], stable_keys: list[str]) -> bool:
    """True if row agrees with pattern on every non-empty stable field."""
    for k in stable_keys:
        if k not in pattern:
            continue
        ps = str(pattern.get(k, "") or "").strip()
        rs = str(row.get(k, "") or "").strip()
        if ps and rs and ps != rs:
            return False
    return True


def _enrich_pattern_from_rates_table(
    pattern: dict[str, str],
    rates_table: list[dict],
    stable_keys: list[str],
) -> dict[str, str]:
    """
    Fill missing pattern fields from the rate card `rates_table` when a field is **constant** on every
    row that matches this pattern on stable_keys.

    extract_patterns() uses the whole table: a column like Destination Postal has many distinct values
    *globally* (outbound zones + return postal), so it is marked "changing" and omitted from patterns —
    even when it is constant for all outbound rows (e.g. only) or all RETURN rows (e.g. 73110).
    """
    out = dict(pattern)
    if not rates_table:
        return out
    matching = [
        r
        for r in rates_table
        if isinstance(r, dict) and _rate_row_matches_stable_pattern(r, out, stable_keys)
    ]
    if not matching:
        return out
    all_keys: set[str] = set()
    for r in matching:
        all_keys.update(r.keys())
    for k in all_keys:
        if k in ("Lane #", "Lane"):
            continue
        vals = {
            str(r.get(k, "") or "").strip()
            for r in matching
            if r.get(k) is not None and str(r.get(k)).strip()
        }
        vals = {v for v in vals if v}
        if len(vals) != 1:
            continue
        v = next(iter(vals))
        if k not in out or not str(out.get(k, "") or "").strip():
            out[k] = v
    return out


def _row_has_rate_amount(dr: dict) -> bool:
    """True if row looks like a tariff line: common currency symbols or EU-style decimal amounts (PLN, etc.)."""
    for v in dr.values():
        if v is None:
            continue
        s = str(v)
        if any(sym in s for sym in ("€", "$", "£", "¥")):
            return True
        low = s.lower()
        if "kč" in low:
            return True
        if "zł" in low:
            return True
        if re.search(r"\b(PLN|EUR|USD|GBP|CHF|CZK|HUF)\b", s, re.I):
            return True
        # e.g. 12,34 / 12.34 — typical in PL tariffs without € in every cell
        if re.search(r"\d+[.,]\d{2}\b", s):
            return True
    return False


# Column label (from tariff header row) → pattern key(s) it can fill (order = try first)
_LABEL_TO_PATTERN_KEYS: list[tuple[re.Pattern, tuple[str, ...]]] = [
    # "Zip 2" header is handled separately → Destination City (origin) / Origin City (return), not postal
    (re.compile(r"origin\s+postal|from\s*postal|frompostalcode", re.I), ("Origin Postal Code",)),
    (re.compile(r"destination\s+postal|to\s*postal|topostalcode", re.I), ("Destination Postal Code",)),
    (re.compile(r"origin\s+city", re.I), ("Origin City",)),
    (re.compile(r"destination\s+city", re.I), ("Destination City",)),
    (re.compile(r"area\s*code", re.I), ("Area Code",)),
    (re.compile(r"^zip\s*$|postal\s*code", re.I), ("Origin Postal Code", "Destination Postal Code")),
]

# Keys we always try to fill from tariff headers (rate_card "patterns" may omit these as "changing")
_EXTRA_PATTERN_FILL_KEYS = (
    "Origin Postal Code",
    "Destination Postal Code",
    "Origin City",
    "Destination City",
    "Area Code",
)

# Only fill these from tariff / Zip 2 if the key is non-empty on the extracted pattern (before enrichment).
_CITY_KEYS_FOR_TARIFF_FILL = frozenset({"Origin City", "Destination City"})


def _city_keys_declared_in_rate_card(pattern: dict[str, str]) -> frozenset[str]:
    """
    City keys that are **populated** on the extracted rate-card pattern (before enrichment).

    Each pattern dict includes every *stable* column as a key, so ``Origin City`` and
    ``Destination City`` can both be present with ``""``. Only keys with a non-empty value
    count as declared for that lane; the other side must stay blank (no enrichment/tariff fill).
    """
    out: set[str] = set()
    for k in _CITY_KEYS_FOR_TARIFF_FILL:
        if k not in pattern:
            continue
        if str(pattern.get(k, "") or "").strip():
            out.add(k)
    return frozenset(out)


def _strip_undeclared_city_keys(
    expanded: dict[str, str],
    city_keys_from_rate_card: frozenset[str] | None,
) -> dict[str, str]:
    """
    Clear Origin/Destination City when not on the **extracted** pattern (before enrichment).

    :func:`_enrich_pattern_from_rates_table` may add constant cities from the rate-card table;
    those would otherwise appear in the matrix even when tariff fill is gated off.
    """
    if city_keys_from_rate_card is None:
        return expanded
    out = dict(expanded)
    for k in _CITY_KEYS_FOR_TARIFF_FILL:
        if k not in city_keys_from_rate_card:
            out[k] = ""
    return out


def _column_sort_key(name: str):
    m = re.match(r"^Column(\d+)$", str(name))
    return (0, int(m.group(1))) if m else (1, name)


def _sorted_column_keys(row: dict) -> list[str]:
    return sorted(row.keys(), key=_column_sort_key)


def _primary_column_key(row: dict) -> str | None:
    """
    Leftmost ColumnN that has a value. After _clean_data(), empty leading columns are
    removed from JSON, so 'Zip 2' / bands may live in Column1+ only — we must not assume Column0.
    """
    if not row:
        return None
    for k in _sorted_column_keys(row):
        if not re.match(r"^Column\d+$", k):
            continue
        v = row.get(k)
        if v is not None and str(v).strip():
            return k
    return None


def _primary_value(row: dict) -> str:
    k = _primary_column_key(row)
    if not k:
        return ""
    return str(row.get(k) or "").strip()


# Substrings that identify a *column role* when they appear in a cell (header row defines ColumnK -> label)
_COLUMN_ROLE_SUBSTRINGS = (
    "zip 2",
    "destination postal",
    "origin postal",
    "area code",
    "postal code",
    "postcode",
    "plz",
    "category",
    "service",
    "carrier",
    "tariff sheet",
)


def _cell_is_column_role_label(text: str) -> bool:
    if not text or not str(text).strip():
        return False
    s = str(text).strip().lower()
    for sub in _COLUMN_ROLE_SUBSTRINGS:
        if sub in s:
            return True
    # lone "zip" as word (avoid matching random numbers)
    if re.search(r"\bzip\b", s) and "zip 2" not in s and len(s) <= 12:
        return True
    return False


def _any_cell_is_column_role_label(row: dict) -> bool:
    """True if any cell names a column role (e.g. 'Destination Postal Code', 'Zip 2')."""
    if not row:
        return False
    for v in row.values():
        if v is not None and _cell_is_column_role_label(str(v)):
            return True
    return False


def _count_weight_band_header_cells(row: dict) -> int:
    """How many cells look like weight-band headers (0-10, 11-20, > 10000), not € rates."""
    n = 0
    for v in row.values():
        if v is None:
            continue
        t = str(v).strip()
        if not t or "€" in t or "$" in t:
            continue
        if re.match(r"^\d+\s*[-–]\s*\d+$", t):
            n += 1
        elif re.match(r"^>\s*\d", t.replace(" ", "")):
            n += 1
    return n


def _is_tariff_header_row(row: dict) -> bool:
    """
    Header row: JSON keys are Column0, Column1, … and *values* in that row are the
    human labels ('Zip 2', 'Destination Postal Code', weight bands '0-10', …).
    We scan all cells — the lane column is not always Column0 (e.g. RDE6 uses Column1).
    """
    if not row:
        return False
    if _any_cell_is_column_role_label(row):
        return True
    # Row that is mostly weight-band headers (0-10, 11-20, …) without € — typical matrix header
    if _count_weight_band_header_cells(row) >= 5:
        return True
    return False


def _is_data_row(row: dict) -> bool:
    """Data row: not a title/header; first populated column has band or postal value, or row has rates."""
    if _get_table_title_from_row(row):
        return False
    if _is_tariff_header_row(row):
        return False
    pv = _primary_value(row)
    if not pv:
        # e.g. leading empty columns — use any cell
        for v in row.values():
            if v is None:
                continue
            s = str(v).strip()
            if "€" in s or "$" in s:
                return True
            if re.search(r"\d", s) and not _cell_is_column_role_label(s):
                return True
        return False
    if _cell_is_column_role_label(pv):
        return False
    if "price per shipment" in pv.lower() or "price per kg" in pv.lower():
        return False
    if bool(re.search(r"\d", pv)):
        return True
    return any("€" in str(v) or "$" in str(v) for v in row.values() if v is not None)


def _row_looks_like_tariff_data_fallback(row: dict) -> bool:
    """If strict rules found no rows, accept rows with band-like values in any column."""
    if _get_table_title_from_row(row):
        return False
    if _is_tariff_header_row(row):
        return False

    def bandish(s: str) -> bool:
        s = str(s).strip()
        if not s:
            return False
        if re.search(r"\d+\s*[-–]\s*\d+", s):
            return True
        if re.match(r"^\d{2,}$", s):
            return True
        return bool(re.search(r"\d", s))

    for v in row.values():
        if v is not None and bandish(str(v)):
            return True
    return False


def _label_base_for_match(label: str) -> str:
    """Strip duplicate disambiguation suffix ' (Column25)' for regex matching."""
    return re.sub(r"\s*\(Column\d+\)\s*$", "", label or "").strip()


def _unique_column_labels(header_row: dict) -> dict[str, str]:
    """
    Map ColumnK -> header cell value. If the same label appears twice (side-by-side matrices),
    disambiguate: 'Destination Postal Code (Column1)', 'Destination Postal Code (Column25)'.
    """
    raw: dict[str, str] = {}
    for ck in _sorted_column_keys(header_row):
        lab = header_row.get(ck)
        if lab is not None and str(lab).strip():
            raw[ck] = str(lab).strip()
    counts: dict[str, int] = {}
    for lab in raw.values():
        counts[lab] = counts.get(lab, 0) + 1
    out: dict[str, str] = {}
    for ck, lab in raw.items():
        out[ck] = f"{lab} ({ck})" if counts.get(lab, 0) > 1 else lab
    return out


def _find_tariff_header_row_index(table_rows: list[dict]) -> int | None:
    for i, row in enumerate(table_rows):
        if _is_tariff_header_row(row):
            return i
    return None


def _cell_looks_like_rate_tier_banner(s: str) -> bool:
    """Row above bands: 'Rates Per Shipment', 'Rate per kg', 'Price per 100kg', 'FTL', etc."""
    if not s or not str(s).strip():
        return False
    t = str(s).strip()
    if _is_weight_band_label(t):
        return False
    if re.match(r"^\d+$", t):
        return False
    tl = t.lower()
    if re.search(r"per\s*[\d.,]+\s*kg", tl):
        return True
    if "price per" in tl:
        return True
    return any(
        k in tl
        for k in (
            "shipment",
            "per kg",
            "rate per",
            "/kg",
            "ftl",
            "per unit",
            "flat",
        )
    )


def _row_looks_like_rate_tier_meta(row: dict) -> bool:
    blob = " ".join(str(v) for v in row.values() if v is not None and str(v).strip()).lower()
    if "shipment" in blob or "per kg" in blob or "rate per" in blob or "ftl" in blob:
        return True
    if "price per" in blob:
        return True
    if re.search(r"per\s*[\d.,]+\s*kg", blob):
        return True
    return False


def _forward_fill_rate_tier_text_by_column(meta_row: dict, header_row: dict) -> dict[str, str]:
    """ColumnK -> tier label from cells like 'Rates Per Shipment' (forward-filled until next banner)."""
    # Tier row may be sparse (only Column2, Column9, …); walk **all** header keys so gaps inherit.
    keys_union = set(_sorted_column_keys(meta_row)) | set(_sorted_column_keys(header_row))
    keys = sorted(keys_union, key=_column_sort_key)
    current: str | None = None
    out: dict[str, str] = {}
    for ck in keys:
        v = meta_row.get(ck)
        if v is not None and str(v).strip():
            s = str(v).strip()
            if _cell_looks_like_rate_tier_banner(s):
                current = s
        if current:
            out[ck] = current
    return out


def _tier_text_to_measure_label(tier_text: str) -> str:
    """
    Display under weight bracket: Flat (per shipment), p/unit (per kg, no step), p/X unit (per X kg).
    Matches e.g. 'Rate per kg', 'Price per 100kg', 'per 10 kg'.
    """
    if not tier_text:
        return ""
    tl = tier_text.lower()
    if "shipment" in tl:
        return "Flat"
    if re.search(r"\bftl\b", tl) and "per" not in tl:
        return "Flat"
    # Per X kg (X explicit) before generic 'per kg', e.g. Price per 100kg → p/100 unit
    m = re.search(r"per\s*([\d.,]+)\s*kg", tier_text, re.I)
    if m:
        x = m.group(1).replace(",", ".").strip()
        return f"p/{x} unit"
    if "per kg" in tl or re.search(r"rate\s*per\s*kg", tier_text, re.I) or re.search(
        r"rate\s*per\s*/?\s*kg", tl
    ):
        return "p/unit"
    return ""


def extract_rate_measure_labels_by_display_name(table_rows: list[dict]) -> dict[str, str]:
    """
    Map tariff **display** column label (e.g. '<10,99') -> measure row text ('Flat', 'p/unit', …)
    using the row **above** the band header when present (FIEGE: Rates Per Shipment / Rate per kg).
    """
    hi = _find_tariff_header_row_index(table_rows)
    if hi is None or hi < 1:
        return {}
    meta = table_rows[hi - 1]
    if not _row_looks_like_rate_tier_meta(meta):
        return {}
    header_row = table_rows[hi]
    tier_by_ck = _forward_fill_rate_tier_text_by_column(meta, header_row)
    labels = _unique_column_labels(header_row)
    out: dict[str, str] = {}
    for ck, lab in labels.items():
        tier = tier_by_ck.get(ck)
        if not tier:
            continue
        mlabel = _tier_text_to_measure_label(tier)
        if not mlabel:
            continue
        base = _label_base_for_match(lab)
        out[base] = mlabel
        out[lab] = mlabel
    return out


def _measure_list_for_rate_block_columns(
    rate_block_cols: list[str], measure_by_label: dict[str, str]
) -> list[str]:
    row: list[str] = []
    for col in rate_block_cols:
        if _is_currency_column_name(col):
            row.append("")
            continue
        if _is_geq_synthetic_bracket_label(col):
            row.append("Flat")
            continue
        base = _label_base_for_match(col)
        v = measure_by_label.get(col) or measure_by_label.get(base) or ""
        row.append(v)
    return row


def _parse_header_and_data(table_rows: list[dict]) -> tuple[dict[str, str], list[dict]]:
    """
    From one table block, find header row (Zip 2 / weight bands) and following data rows.
    Returns (column_key -> display label, list of data rows as full dicts).
    """
    if not table_rows:
        return {}, []
    header_idx = _find_tariff_header_row_index(table_rows)
    if header_idx is None:
        data = [r for r in table_rows if _is_data_row(r)]
        if not data:
            data = [r for r in table_rows if _row_looks_like_tariff_data_fallback(r)]
        return {}, data
    header_row = table_rows[header_idx]
    labels = _unique_column_labels(header_row)
    data_rows = []
    for row in table_rows[header_idx + 1 :]:
        if _get_table_title_from_row(row):
            break
        if _is_tariff_header_row(row):
            break
        if _is_data_row(row):
            data_rows.append(row)
    if not data_rows:
        for row in table_rows[header_idx + 1 :]:
            if _get_table_title_from_row(row):
                break
            if _is_tariff_header_row(row):
                break
            if _row_looks_like_tariff_data_fallback(row):
                data_rows.append(row)
    return labels, data_rows


def _format_pattern_lane(p: dict) -> str:
    """Short string for user prompt, e.g. 'CZ, Bor u Tachova → AT, '."""
    oc = (p.get("Origin Country") or "").strip()
    ocity = (p.get("Origin City") or "").strip()
    dc = (p.get("Destination Country") or "").strip()
    dcity = (p.get("Destination City") or "").strip()
    left = ", ".join(x for x in [oc, ocity] if x) or "—"
    right = ", ".join(x for x in [dc, dcity] if x) or "—"
    return f"{left} → {right}"


# Order for human-readable pattern dumps in prompts (then any remaining keys).
_PATTERN_DETAIL_KEY_ORDER: tuple[str, ...] = (
    "Carrier",
    "Tariff sheet",
    "Origin Country",
    "Origin City",
    "Origin Postal Code",
    "Destination Country",
    "Destination City",
    "Destination Postal Code",
    "Service",
    "Lane #",
    "Lane",
)


def _format_pattern_detail_block(p: dict, *, indent: str = "    ") -> str:
    """
    Multi-line listing of non-empty pattern fields for interactive prompts.
    Known keys first, then any other keys (alphabetically), excluding private keys.
    """
    lines: list[str] = []
    seen: set[str] = set()
    for k in _PATTERN_DETAIL_KEY_ORDER:
        if k not in p:
            continue
        seen.add(k)
        v = p.get(k)
        if v is None:
            continue
        s = str(v).strip()
        if not s:
            continue
        lines.append(f"{indent}{k}: {s}")
    for k in sorted(p.keys()):
        if k in seen or k.startswith("_"):
            continue
        v = p.get(k)
        if v is None:
            continue
        s = str(v).strip()
        if not s:
            continue
        lines.append(f"{indent}{k}: {s}")
    if not lines:
        return f"{indent}(no fields)"
    return "\n".join(lines)


def _format_pattern_one_line_summary(p: dict) -> str:
    """Compact line for summaries: carrier, tariff sheet label, lane, service."""
    bits: list[str] = []
    c = (p.get("Carrier") or "").strip()
    ts = (p.get("Tariff sheet") or "").strip()
    if c:
        bits.append(c)
    if ts:
        bits.append(f"[{ts}]")
    lane = _format_pattern_lane(p)
    if lane and lane != "— → —":
        bits.append(lane)
    svc = (p.get("Service") or "").strip()
    if svc:
        bits.append(f"Service {svc}")
    return " · ".join(bits) if bits else _format_pattern_lane(p)


def _print_pattern_catalog(patterns: list[dict], *, intro: str) -> None:
    """Print each pattern with summary + full field list (for origin/return / flat-zero prompts)."""
    print(intro)
    for i, p in enumerate(patterns, 1):
        role = _infer_pattern_lane_role(p)
        print(f"\n  Pattern {i} of {len(patterns)} — lane type for city/zip fill: {role}")
        print(f"    Summary: {_format_pattern_one_line_summary(p)}")
        print("    All fields from rate card:")
        print(_format_pattern_detail_block(p))


def _summarize_table_block_line(table_index_1based: int, tb: list[dict]) -> str:
    """Multi-line block for prompts: index, row count, return/outbound hint, title + row peek."""
    n = len(tb)
    head = f"  Table {table_index_1based}: {n} rows"
    if not tb:
        return f"{head}\n      (empty block)"
    blob_raw = _table_title_blob(tb[0])
    blob_display = (blob_raw[:220] + "…") if len(blob_raw) > 220 else blob_raw
    if not blob_display.strip():
        blob_display = "(no recognised title row — preamble or unnamed block)"
    if _title_indicates_return_lane(blob_raw):
        lane_guess = "return lane (title mentions return)"
    else:
        lane_guess = "outbound / standard (title does not indicate return)"
    peek = ""
    if tb[0]:
        sample_vals = [
            str(v).strip()
            for _, v in sorted(tb[0].items(), key=lambda x: _column_sort_key(x[0]))
            if v is not None and str(v).strip()
        ]
        if sample_vals:
            peek = " | ".join(sample_vals[:5])
            if len(peek) > 140:
                peek = peek[:137] + "…"
    lines = [
        f"{head} — {lane_guess}",
        f"    Title / first row text: {blob_display!r}",
    ]
    if peek:
        lines.append(f"    First-row values (sample): {peek}")
    return "\n".join(lines)


def _infer_pattern_lane_role(pat: dict[str, str]) -> str:
    """Outbound vs return row generation for :func:`_fill_pattern_gaps` (Zip 2 / city fill)."""
    svc = str(pat.get("Service") or "")
    if "RETURN" in svc.upper():
        return "return"
    return "origin"


def _prompt_pattern_to_tariff_table_indices(patterns: list[dict], tables: list[list[dict]]) -> list[int]:
    """
    Ask which merged tariff block (1..len(tables)) applies to each rate-card pattern.
    Returns 0-based indices into ``tables`` (same order as ``patterns``).
    """
    if not patterns or not tables:
        return []
    _ui_heading("Step 3 — Link each rate-card pattern to a tariff table")
    print(
        "Below are the tariff table blocks found after merging your workbook(s).\n"
        "For each lane pattern from the rate card, type which table number (1 to N) supplies its rates.\n"
        "Whether a row is treated as outbound or return for city/zip filling follows the Service field "
        '(e.g. names containing "RETURN" use the return-lane rules).\n'
    )
    print("Tariff table blocks:")
    for ti, tb in enumerate(tables, 1):
        print(_summarize_table_block_line(ti, tb))
    out: list[int] = []
    for pi, p in enumerate(patterns):
        role = _infer_pattern_lane_role(p)
        print(f"\n  Pattern {pi + 1} of {len(patterns)} — lane type: {role}")
        print(f"      Summary: {_format_pattern_one_line_summary(p)}")
        print("      Fields from rate card:")
        print(_format_pattern_detail_block(p, indent="        "))
        while True:
            raw = input(
                f"\n  Which table number for pattern {pi + 1}? (1–{len(tables)}): "
            ).strip()
            try:
                tn = int(raw)
                if 1 <= tn <= len(tables):
                    out.append(tn - 1)
                    break
            except ValueError:
                pass
            print(f"  Please enter a whole number from 1 to {len(tables)}.")
    return out


def _table_title_blob(first_row: dict | None) -> str:
    """All text from first row of a table block, lowercased (for 'return' detection)."""
    if not first_row:
        return ""
    parts = [
        str(v).strip()
        for _, v in sorted(first_row.items(), key=lambda x: _column_sort_key(x[0]))
        if v is not None and str(v).strip()
    ]
    return " ".join(parts).lower()


def _title_indicates_return_lane(blob: str) -> bool:
    """True if table title describes return / reverse freight (incl. typo 'Rerurn')."""
    if not blob:
        return False
    if "return" in blob or "returns" in blob:
        return True
    if "rerurn" in blob:  # common typo in source files
        return True
    return False


def _tables_starting_with_real_title(tables: list[list[dict]]) -> list[list[dict]]:
    """
    _split_into_tables puts rows *before* the first title row into block[0] (preamble: fuel,
    diesel, current share, etc.). Those blocks do not start with a row that rate_to_json treats
    as a table title — exclude them when choosing origin/return freight tables.
    """
    out = [tb for tb in tables if tb and _get_table_title_from_row(tb[0])]
    return out if out else tables


def pick_origin_return_tables(tables: list[list[dict]]) -> tuple[list[dict], list[dict]]:
    """
    Origin table: first freight block whose title does not indicate return (e.g. outbound).
    Return table: first block whose title indicates return.

    Tariff rows come from the same pipeline as rate_to_json.py (sheet_to_json + _clean_data),
    not from running clean_rate_json.py separately — same rules.

    Preamble-only blocks (no title on first row) are skipped — see _tables_starting_with_real_title.
    """
    if not tables:
        return [], []
    use = _tables_starting_with_real_title(tables)
    if _DEBUG and len(use) < len(tables):
        _debug_line(
            f"Skipped {len(tables) - len(use)} block(s) without a real table title on row 1 "
            f"(preamble before 'New Freight Rate' / 'Rates RDE' / …).",
            0,
        )

    if len(use) == 1:
        t0 = use[0]
        return t0, t0

    origin_tb = None
    for tb in use:
        if tb and not _title_indicates_return_lane(_table_title_blob(tb[0])):
            origin_tb = tb
            break
    if origin_tb is None:
        origin_tb = use[0]

    return_tb = None
    for tb in use:
        if tb and _title_indicates_return_lane(_table_title_blob(tb[0])):
            return_tb = tb
            break
    if return_tb is None:
        return_tb = use[1] if len(use) > 1 else use[0]

    return origin_tb, return_tb


def resplit_tables_by_titles(tables: list[list[dict]]) -> list[list[dict]]:
    """
    If the whole sheet landed as one block, split again on every row that is a table title
    (e.g. second 'New Freight Rate Returns…' row). _split_into_tables sometimes yields one
    block depending on workbook shape; this recovers outbound vs return tables.
    """
    if len(tables) >= 2:
        return tables
    if not tables or not tables[0]:
        return tables
    block = tables[0]
    idxs = [i for i, row in enumerate(block) if _get_table_title_from_row(row)]
    if len(idxs) <= 1:
        return tables
    out: list[list[dict]] = []
    for j, start in enumerate(idxs):
        end = idxs[j + 1] if j + 1 < len(idxs) else len(block)
        out.append(block[start:end])
    return out


def _expand_pattern_for_fill(
    pattern: dict[str, str],
    key_order: list[str],
    *,
    changing_keys: list[str] | None = None,
) -> dict[str, str]:
    """Ensure postal/city keys exist so Zip 2 / headers can fill them (stable_keys may omit them)."""
    out = dict(pattern)
    for k in key_order:
        out.setdefault(k, "")
    for k in _EXTRA_PATTERN_FILL_KEYS:
        out.setdefault(k, "")
    if changing_keys:
        for k in changing_keys:
            if k:
                out.setdefault(k, "")
    return out


def _label_is_zip2_lane(base: str) -> bool:
    b = (base or "").strip().lower()
    return b in ("zip 2", "zip2") or re.match(r"^zip\s*2$", b) is not None


def _zip_lane_column_keys(col_labels: dict[str, str]) -> set[str]:
    keys: set[str] = set()
    for ck, lab in col_labels.items():
        if _label_is_zip2_lane(_label_base_for_match(lab)):
            keys.add(ck)
    return keys


def _is_informational_no_rate_row(dr: dict) -> bool:
    """Notes like 'AT returns … increased by 10% against standard rates' — no tariff line."""
    if _row_has_rate_amount(dr):
        return False
    text = " ".join(str(v) for v in dr.values() if v is not None and str(v).strip()).strip()
    if not text:
        return True
    tl = text.lower()
    if any(
        x in tl
        for x in (
            "%",
            "increase",
            "standard rate",
            "against standard",
            "will be increased",
            "both parts",
            "diesel and none",
            "against standard rates",
        )
    ):
        return True
    if len(text) > 100:
        return True
    return False


def _fill_pattern_gaps(
    pattern: dict[str, str],
    tariff_row: dict,
    col_labels: dict[str, str],
    *,
    lane_role: str,
    changing_keys: list[str] | None = None,
    city_keys_from_rate_card: frozenset[str] | None = None,
) -> dict[str, str]:
    """
    Fill empty pattern fields from tariff columns. lane_role is 'origin' or 'return'.

    **Zip 2** column: In many Bridgestone tariffs this is a **postal zone** (see conditional rules
    on FROMPOSTALCODE / TOPOSTALCODE). If *changing_keys* include postal fields, Zip 2 fills those
    (outbound → destination postal first; return → origin postal first). Only when postals are not
    among changing keys do we use the legacy behaviour: Zip 2 → empty **Destination City** (origin)
    or **Origin City** (return).

    If *city_keys_from_rate_card* is set, **Origin City** / **Destination City** are filled from
    tariff or Zip 2 only when that key was present on the extracted rate-card pattern (before
    enrichment). Missing keys stay empty so outbound/return patterns stay one-sided.

    The Zip 2 column is still omitted from the wide tariff part (see _zip_lane_column_keys).
    """
    out = dict(pattern)
    if _DEBUG:
        _debug_line(f"_fill_pattern_gaps(lane_role={lane_role!r})", 0)
        _debug_line(f"Header col_labels ({len(col_labels)}):", 1)
        for ck, lab in list(col_labels.items())[:24]:
            _debug_line(f"  {ck} -> {_label_base_for_match(lab)!r}", 2)
        if len(col_labels) > 24:
            _debug_line(f"  ... +{len(col_labels) - 24} more columns", 2)
        empty_keys = [k for k, v in out.items() if not str(v or "").strip()]
        _debug_line(f"Pattern fields empty before fill: {empty_keys}", 1)
        pv = _primary_value(tariff_row)
        _debug_line(f"Tariff row primary cell (lane/band column): {pv!r}", 1)
        zip_cks = [ck for ck, lab in col_labels.items() if _label_is_zip2_lane(_label_base_for_match(lab))]
        _debug_line(f"Columns detected as Zip 2 lane: {zip_cks}", 1)
    for pk in list(out.keys()):
        if (
            city_keys_from_rate_card is not None
            and pk in _CITY_KEYS_FOR_TARIFF_FILL
            and pk not in city_keys_from_rate_card
        ):
            continue
        if str(out.get(pk) or "").strip():
            continue
        for ck, label in col_labels.items():
            base = _label_base_for_match(label or "")
            for rx, pkeys in _LABEL_TO_PATTERN_KEYS:
                if pk not in pkeys:
                    continue
                if rx.search(base):
                    val = tariff_row.get(ck)
                    if val is not None and str(val).strip():
                        out[pk] = str(val).strip()
                        break
            if str(out.get(pk) or "").strip():
                break
        # Legacy: Zip in header text (not necessarily "Zip 2") for postal keys — rare
        if not str(out.get(pk) or "").strip() and "postal" in pk.lower():
            for ck, label in col_labels.items():
                base = _label_base_for_match(label or "").lower()
                if "zip" not in base and base != "":
                    continue
                if _label_is_zip2_lane(_label_base_for_match(label or "")):
                    continue
                v = tariff_row.get(ck)
                if v is None or not str(v).strip():
                    continue
                vs = str(v).strip()
                if lane_role == "origin" and "origin" in pk.lower() and "destination" not in pk.lower():
                    out[pk] = vs
                    break
                if lane_role == "return" and "destination" in pk.lower():
                    out[pk] = vs
                    break

    # Changing dimensions from rate-card analysis (names vary) ← lane columns left of weight bands (e.g. Area Code)
    if changing_keys:
        lane_cks = _tariff_lane_column_keys(col_labels)
        if _DEBUG:
            _debug_line(f"changing_keys (fill from tariff lane cols): {changing_keys}", 1)
            _debug_line(f"tariff lane column keys (left of first band): {lane_cks}", 1)
        li = 0
        for pk in changing_keys:
            if not pk or pk not in out:
                continue
            if (
                city_keys_from_rate_card is not None
                and pk in _CITY_KEYS_FOR_TARIFF_FILL
                and pk not in city_keys_from_rate_card
            ):
                continue
            if str(out.get(pk) or "").strip():
                continue
            while li < len(lane_cks):
                ck = lane_cks[li]
                li += 1
                v = tariff_row.get(ck)
                if v is not None and str(v).strip():
                    out[pk] = str(v).strip()
                    break

    # Zip 2 column: postal zones (FROMPOSTAL/TOPOSTAL in rate card) vs legacy → City
    dest_postal_keys = [
        k for k in (changing_keys or []) if k and "destination" in k.lower() and "postal" in k.lower()
    ]
    orig_postal_keys = [
        k for k in (changing_keys or []) if k and "origin" in k.lower() and "postal" in k.lower()
    ]
    postal_in_changing = bool(dest_postal_keys or orig_postal_keys)

    for ck, label in col_labels.items():
        base = _label_base_for_match(label)
        if not _label_is_zip2_lane(base):
            continue
        v = tariff_row.get(ck)
        if v is None or not str(v).strip():
            continue
        vs = str(v).strip()
        if postal_in_changing:
            if lane_role == "origin":
                for pk in dest_postal_keys:
                    if pk in out and not str(out.get(pk) or "").strip():
                        out[pk] = vs
                        break
                else:
                    for pk in orig_postal_keys:
                        if pk in out and not str(out.get(pk) or "").strip():
                            out[pk] = vs
                            break
            else:
                for pk in orig_postal_keys:
                    if pk in out and not str(out.get(pk) or "").strip():
                        out[pk] = vs
                        break
                else:
                    for pk in dest_postal_keys:
                        if pk in out and not str(out.get(pk) or "").strip():
                            out[pk] = vs
                            break
            continue
        if lane_role == "origin" and not str(out.get("Destination City") or "").strip():
            if city_keys_from_rate_card is None or "Destination City" in city_keys_from_rate_card:
                out["Destination City"] = vs
        elif lane_role == "return" and not str(out.get("Origin City") or "").strip():
            if city_keys_from_rate_card is None or "Origin City" in city_keys_from_rate_card:
                out["Origin City"] = vs
    if _DEBUG:
        _debug_line(
            "Result: "
            f"Origin City={out.get('Origin City')!r}, "
            f"Destination City={out.get('Destination City')!r}, "
            f"Origin Postal={out.get('Origin Postal Code')!r}, "
            f"Dest Postal={out.get('Destination Postal Code')!r}",
            1,
        )
        if postal_in_changing:
            _debug_line(
                "Zip 2 applied to postal keys from changing_keys (not City); "
                f"dest_postal_keys={dest_postal_keys}, orig_postal_keys={orig_postal_keys}.",
                1,
            )
        elif not str(out.get("Destination City") or "").strip() and lane_role == "origin":
            _debug_line(
                "NOTE: Destination City still empty — Zip 2 → City (legacy) on origin lane if Zip 2 present.",
                1,
            )
        elif not str(out.get("Origin City") or "").strip() and lane_role == "return":
            _debug_line(
                "NOTE: Origin City still empty — Zip 2 → City (legacy) on return lane if Zip 2 present.",
                1,
            )
    return out


def _all_pattern_output_keys(pattern_key_order: list[str]) -> list[str]:
    out: list[str] = []
    for k in list(pattern_key_order) + list(_EXTRA_PATTERN_FILL_KEYS):
        if k not in out:
            out.append(k)
    return out


_EXCEL_OUTPUT_INTERNAL_DROP = frozenset(
    {"_pattern_role", "_pattern_number", "_pattern_index", "_tariff_table_block"}
)


def _rate_card_lane_column_order(rates_table: list) -> list[str]:
    """Column order for the lane block as in the previous-rate workbook (rates table row keys)."""
    if not rates_table or not isinstance(rates_table[0], dict):
        return []
    return list(rates_table[0].keys())


def _excel_shipment_column_prefix(
    df: pd.DataFrame,
    rates_table: list,
    key_order: list[str],
) -> list[str]:
    """Shipment columns: match previous rate card order and presence; omit internal keys."""
    lane = _rate_card_lane_column_order(rates_table)
    if lane:
        return [c for c in lane if c in df.columns and c not in _EXCEL_OUTPUT_INTERNAL_DROP]
    return [
        c
        for c in _all_pattern_output_keys(key_order)
        if c in df.columns and c not in _EXCEL_OUTPUT_INTERNAL_DROP
    ]


def _is_generic_excel_sheet_tab_name(name: str) -> bool:
    """True for default tab names like Sheet1, Sheet 2 (not real business sheet titles)."""
    s = (name or "").strip()
    if not s:
        return True
    return bool(re.match(r"^sheet\s*\d+\s*$", s, re.I))


def _rate_card_has_tariff_sheet_column(rates_table: list) -> bool:
    if not rates_table or not isinstance(rates_table[0], dict):
        return False
    for k in rates_table[0].keys():
        if str(k).strip().lower() == "tariff sheet":
            return True
    return False


def _find_tariff_sheet_dataframe_column(df: pd.DataFrame) -> str | None:
    for c in df.columns:
        if str(c).strip().lower() == "tariff sheet":
            return str(c)
    return None


def _chosen_tariff_sheet_display_value(pairs: list[tuple[Path, str]]) -> str | None:
    """Sheet name(s) to show when all chosen tabs have non-generic names; else None."""
    if not pairs:
        return None
    names = [sn for _, sn in pairs]
    if any(_is_generic_excel_sheet_tab_name(n) for n in names):
        return None
    if len(names) == 1:
        return names[0]
    out: list[str] = []
    for n in names:
        if n not in out:
            out.append(n)
    return " + ".join(out)


def _apply_chosen_tariff_sheet_to_matrix(
    df: pd.DataFrame,
    rates_table: list,
    pairs: list[tuple[Path, str]],
) -> pd.DataFrame:
    """
    If the previous rate card has a **Tariff sheet** lane column and the user picked non-generic
    Excel tab name(s), set every row in that column to the chosen sheet name(s).
    """
    if df.empty or not _rate_card_has_tariff_sheet_column(rates_table):
        return df
    col = _find_tariff_sheet_dataframe_column(df)
    if not col:
        return df
    val = _chosen_tariff_sheet_display_value(pairs)
    if val is None:
        return df
    out = df.copy()
    out[col] = val
    return out


def _build_matrix_rows(
    pattern_key_order: list[str],
    *,
    origin_pattern: dict[str, str],
    return_pattern: dict[str, str] | None,
    origin_table: list[dict],
    return_table: list[dict] | None,
    origin_pattern_idx: int,
    return_pattern_idx: int | None,
    changing_keys: list[str] | None = None,
    flat_zero_pattern: dict[str, str] | None = None,
    flat_zero_pattern_idx: int | None = None,
    flat_zero_base: str | None = None,
    origin_city_keys_from_rate_card: frozenset[str] | None = None,
    return_city_keys_from_rate_card: frozenset[str] | None = None,
    flat_zero_city_keys_from_rate_card: frozenset[str] | None = None,
) -> list[dict]:
    """
    One output row per tariff data row: **origin** block, optional **flat-zero** block, then **return**.

    Flat-zero rows use the flat-zero pattern dict; tariff **Flat** bands → ``0``; **p/unit** (and
    p/X unit) copy from the **zero-base** table (same row index as origin or return data rows).

    City-key frozensets come from :func:`_city_keys_declared_in_rate_card` on the extracted pattern
    (before enrichment); ``None`` restores legacy fill behaviour for cities.
    """
    rows_out: list[dict] = []
    out_keys = _all_pattern_output_keys(pattern_key_order)
    # Tariff lane columns must not overwrite pattern columns when headers match the same display name
    # (e.g. return table "Destination Postal Code" zone 01–31 vs rate card hub 22113).
    pattern_column_names = frozenset(out_keys)

    if _DEBUG:
        _debug_step(10, "Matrix: ORIGIN table → pattern fill")
        _debug_line(f"origin_pattern_idx={origin_pattern_idx} (Excel _pattern_number={origin_pattern_idx + 1})", 0)
        _debug_line(f"Origin pattern dict (from rate card): {origin_pattern}", 0)
        _debug_line(f"pattern_key_order: {pattern_key_order}", 0)
        _debug_line(f"out_keys (pattern columns in output, incl. extra fill keys): {out_keys}", 0)
        _debug_line(f"Origin table block row count: {len(origin_table)}", 0)

    col_labels_o, data_rows_o = _parse_header_and_data(origin_table)
    if _DEBUG:
        _debug_line(f"After _parse_header_and_data: raw data_rows={len(data_rows_o)}", 0)
        if data_rows_o:
            _debug_line(f"First raw data row primary value: {_primary_value(data_rows_o[0])!r}", 1)

    data_rows_o_raw_ct = len(data_rows_o)
    data_rows_o = [
        dr
        for dr in data_rows_o
        if not _is_informational_no_rate_row(dr) and _row_has_rate_amount(dr)
    ]
    if _DEBUG:
        _debug_line(
            f"After filter (currency / rate-like amounts; drop info rows): {len(data_rows_o)} rows "
            f"(dropped {data_rows_o_raw_ct - len(data_rows_o)})",
            0,
        )

    zip_skip_o = _zip_lane_column_keys(col_labels_o)
    expanded_o = _strip_undeclared_city_keys(
        _expand_pattern_for_fill(
            origin_pattern, pattern_key_order, changing_keys=changing_keys
        ),
        origin_city_keys_from_rate_card,
    )
    if _DEBUG:
        _debug_line(f"expanded_o (pattern + empty fill slots): {expanded_o}", 0)
        _debug_line(f"Zip 2 / lane columns omitted from wide tariff part (already in pattern): {zip_skip_o}", 0)
        _debug_line(f"col_labels_o (header → label): {dict(list(col_labels_o.items())[:16])}", 0)

    for i, dr in enumerate(data_rows_o):
        merged: dict[str, str] = {}
        if _DEBUG and i < 3:
            _debug_line(f"--- Origin data row {i + 1} / {len(data_rows_o)} ---", 1)
        filled = _fill_pattern_gaps(
            expanded_o,
            dr,
            col_labels_o,
            lane_role="origin",
            changing_keys=changing_keys,
            city_keys_from_rate_card=origin_city_keys_from_rate_card,
        )
        if _DEBUG and i < 3:
            _debug_line(
                f"merged pattern fields: Dest City={filled.get('Destination City')!r} "
                f"(Zip 2 → Dest City on origin lane), "
                f"Origin Postal={filled.get('Origin Postal Code')!r}, Dest Postal={filled.get('Destination Postal Code')!r}",
                2,
            )
        for pk in out_keys:
            merged[pk] = str(filled.get(pk, "") or "")
        for ck in _sorted_column_keys(dr):
            if ck in zip_skip_o:
                continue
            label = col_labels_o.get(ck, ck)
            if label in pattern_column_names:
                continue
            merged[label] = dr.get(ck)
        merged["_pattern_role"] = "origin"
        merged["_pattern_index"] = origin_pattern_idx
        merged["_pattern_number"] = origin_pattern_idx + 1
        rows_out.append(merged)

    # --- Flat-zero: same row count / alignment as zero-base (origin or return) tariff table ---
    if (
        flat_zero_pattern is not None
        and flat_zero_pattern_idx is not None
        and flat_zero_base in ("origin", "return")
    ):
        base_tb = origin_table if flat_zero_base == "origin" else (return_table or [])
        if _DEBUG:
            _debug_step(10, "Matrix: FLAT-ZERO pattern (Flat bands → 0; p/unit from zero-base table)")
            _debug_line(f"flat_zero_pattern_idx={flat_zero_pattern_idx} (_pattern_number={flat_zero_pattern_idx + 1})", 0)
            _debug_line(f"flat_zero_base={flat_zero_base!r} (tariff columns from that table)", 0)
            _debug_line(f"Flat-zero pattern dict: {flat_zero_pattern}", 0)
            _debug_line(f"Zero-base table row count: {len(base_tb)}", 0)
        if base_tb:
            col_labels_z, data_rows_z = _parse_header_and_data(base_tb)
            data_rows_z_raw = len(data_rows_z)
            data_rows_z = [
                dr
                for dr in data_rows_z
                if not _is_informational_no_rate_row(dr) and _row_has_rate_amount(dr)
            ]
            if _DEBUG:
                _debug_line(
                    f"Flat-zero: after filter {len(data_rows_z)} rows (dropped {data_rows_z_raw - len(data_rows_z)})",
                    0,
                )
            measure_by_z = extract_rate_measure_labels_by_display_name(base_tb)
            zip_skip_z = _zip_lane_column_keys(col_labels_z)
            expanded_z = _strip_undeclared_city_keys(
                _expand_pattern_for_fill(
                    flat_zero_pattern, pattern_key_order, changing_keys=changing_keys
                ),
                flat_zero_city_keys_from_rate_card,
            )
            lane_z = "origin" if flat_zero_base == "origin" else "return"
            for i, dr in enumerate(data_rows_z):
                merged: dict[str, str | int | float | None] = {}
                if _DEBUG and i < 3:
                    _debug_line(f"--- Flat-zero data row {i + 1} / {len(data_rows_z)} ---", 1)
                filled = _fill_pattern_gaps(
                    expanded_z,
                    dr,
                    col_labels_z,
                    lane_role=lane_z,
                    changing_keys=changing_keys,
                    city_keys_from_rate_card=flat_zero_city_keys_from_rate_card,
                )
                for pk in out_keys:
                    merged[pk] = str(filled.get(pk, "") or "")
                for ck in _sorted_column_keys(dr):
                    if ck in zip_skip_z:
                        continue
                    label = col_labels_z.get(ck, ck)
                    if label in pattern_column_names:
                        continue
                    if _label_is_flat_zero_tariff_band(label, measure_by_z):
                        merged[label] = 0
                    else:
                        merged[label] = dr.get(ck)
                merged["_pattern_role"] = "flat_zero"
                merged["_pattern_index"] = flat_zero_pattern_idx
                merged["_pattern_number"] = flat_zero_pattern_idx + 1
                rows_out.append(merged)

    if return_pattern is not None and return_table is not None and return_pattern_idx is not None:
        if _DEBUG:
            _debug_step(11, "Matrix: RETURN table → pattern fill")
            _debug_line(f"return_pattern_idx={return_pattern_idx} (_pattern_number={return_pattern_idx + 1})", 0)
            _debug_line(f"Return pattern dict: {return_pattern}", 0)
            _debug_line(f"Return table block row count: {len(return_table)}", 0)

        col_labels_r, data_rows_r = _parse_header_and_data(return_table)
        if _DEBUG:
            _debug_line(f"Return: raw data_rows={len(data_rows_r)}", 0)
        data_rows_r_raw = len(data_rows_r)
        data_rows_r = [
            dr
            for dr in data_rows_r
            if not _is_informational_no_rate_row(dr) and _row_has_rate_amount(dr)
        ]
        if _DEBUG:
            _debug_line(
                f"Return: after filter {len(data_rows_r)} rows (dropped {data_rows_r_raw - len(data_rows_r)})",
                0,
            )
        zip_skip_r = _zip_lane_column_keys(col_labels_r)
        expanded_r = _strip_undeclared_city_keys(
            _expand_pattern_for_fill(
                return_pattern, pattern_key_order, changing_keys=changing_keys
            ),
            return_city_keys_from_rate_card,
        )
        if _DEBUG:
            _debug_line(f"expanded_r: {expanded_r}", 0)
            _debug_line(f"col_labels_r sample: {dict(list(col_labels_r.items())[:12])}", 0)

        for i, dr in enumerate(data_rows_r):
            merged = {}
            if _DEBUG and i < 3:
                _debug_line(f"--- Return data row {i + 1} / {len(data_rows_r)} ---", 1)
            filled = _fill_pattern_gaps(
                expanded_r,
                dr,
                col_labels_r,
                lane_role="return",
                changing_keys=changing_keys,
                city_keys_from_rate_card=return_city_keys_from_rate_card,
            )
            if _DEBUG and i < 3:
                _debug_line(
                    f"merged pattern fields: Origin City={filled.get('Origin City')!r} "
                    f"(Zip 2 → Origin City on return lane), "
                    f"Origin Postal={filled.get('Origin Postal Code')!r}, Dest Postal={filled.get('Destination Postal Code')!r}",
                    2,
                )
            for pk in out_keys:
                merged[pk] = str(filled.get(pk, "") or "")
            for ck in _sorted_column_keys(dr):
                if ck in zip_skip_r:
                    continue
                label = col_labels_r.get(ck, ck)
                if label in pattern_column_names:
                    continue
                merged[label] = dr.get(ck)
            merged["_pattern_role"] = "return"
            merged["_pattern_index"] = return_pattern_idx
            merged["_pattern_number"] = return_pattern_idx + 1
            rows_out.append(merged)

    return rows_out


def _build_matrix_rows_multi(
    pattern_key_order: list[str],
    *,
    patterns: list[dict],
    pattern_table_indices: list[int],
    tables: list[list[dict]],
    changing_keys: list[str] | None = None,
    pattern_city_keys_from_rate_card: list[frozenset[str]] | None = None,
) -> list[dict]:
    """
    One matrix section per rate-card pattern, each using its assigned tariff table block.
    Row metadata includes ``_tariff_table_block`` (1-based, same numbering as prompts).

    *pattern_city_keys_from_rate_card* has one frozenset per pattern (from
    :func:`_city_keys_declared_in_rate_card` on the extracted pattern before enrichment).
    """
    rows_out: list[dict] = []
    out_keys = _all_pattern_output_keys(pattern_key_order)
    pattern_column_names = frozenset(out_keys)
    if len(pattern_table_indices) != len(patterns):
        raise ValueError("pattern_table_indices and patterns length mismatch")

    if _DEBUG:
        _debug_step(10, "Matrix: per-pattern tables (multi tariff + many patterns)")
        for pi, p in enumerate(patterns):
            tix = pattern_table_indices[pi]
            _debug_line(
                f"Pattern {pi + 1} → table block {tix + 1} ({_infer_pattern_lane_role(p)} lane)",
                0,
            )

    for pi, pat in enumerate(patterns):
        tidx = pattern_table_indices[pi]
        if tidx < 0 or tidx >= len(tables):
            continue
        table = tables[tidx]
        role = _infer_pattern_lane_role(pat)
        lane_role: str = "return" if role == "return" else "origin"

        col_labels, data_rows = _parse_header_and_data(table)
        data_rows_raw_ct = len(data_rows)
        data_rows = [
            dr
            for dr in data_rows
            if not _is_informational_no_rate_row(dr) and _row_has_rate_amount(dr)
        ]
        if _DEBUG:
            _debug_line(
                f"Pattern {pi + 1}: table {tidx + 1}, after rate filter "
                f"{len(data_rows)} rows (dropped {data_rows_raw_ct - len(data_rows)})",
                0,
            )

        city_keys_card = (
            pattern_city_keys_from_rate_card[pi]
            if pattern_city_keys_from_rate_card is not None
            and pi < len(pattern_city_keys_from_rate_card)
            else None
        )
        zip_skip = _zip_lane_column_keys(col_labels)
        expanded = _strip_undeclared_city_keys(
            _expand_pattern_for_fill(pat, pattern_key_order, changing_keys=changing_keys),
            city_keys_card,
        )

        for i, dr in enumerate(data_rows):
            merged: dict[str, str | int | float | None] = {}
            if _DEBUG and i < 2:
                _debug_line(f"  Pattern {pi + 1} data row {i + 1}/{len(data_rows)}", 1)
            filled = _fill_pattern_gaps(
                expanded,
                dr,
                col_labels,
                lane_role=lane_role,
                changing_keys=changing_keys,
                city_keys_from_rate_card=city_keys_card,
            )
            for pk in out_keys:
                merged[pk] = str(filled.get(pk, "") or "")
            for ck in _sorted_column_keys(dr):
                if ck in zip_skip:
                    continue
                label = col_labels.get(ck, ck)
                if label in pattern_column_names:
                    continue
                merged[label] = dr.get(ck)
            merged["_pattern_role"] = role if role == "return" else "origin"
            merged["_pattern_index"] = pi
            merged["_pattern_number"] = pi + 1
            merged["_tariff_table_block"] = tidx + 1
            rows_out.append(merged)

    return rows_out


def _collect_pattern_key_order(patterns: list[dict]) -> list[str]:
    seen = []
    for p in patterns:
        for k in p:
            if k not in seen:
                seen.append(k)
    return seen


def run_tariff_pipeline(path: Path, sheet_name: str) -> list[dict]:
    """Same as :func:`rate_to_json.export_tariff_sheet_rows` for one tab."""
    return export_tariff_sheet_rows(path, sheet_name)


def _save_processing_tariff_jsons(
    output_base: str,
    clean_rows: list[dict],
) -> None:
    """
    Writes ``{output_base}.json`` and ``{output_base}_tariff_pipeline.json`` — both match
    ``rate_to_json.py`` output (keyword table selection + clean) and the matrix input.
    """
    PROCESSING_DIR.mkdir(parents=True, exist_ok=True)
    for suffix in (f"{output_base}.json", f"{output_base}_tariff_pipeline.json"):
        with open(PROCESSING_DIR / suffix, "w", encoding="utf-8") as f:
            json.dump(clean_rows, f, indent=2, ensure_ascii=False)


def main():
    _init_debug_from_env()
    # Suppress noisy openpyxl default-style warning when reading xlsx.
    warnings.filterwarnings(
        "ignore",
        message="Workbook contains no default style",
        category=UserWarning,
    )

    # 1) Tariff: input folder (rate_to_json)
    if not TARIFF_INPUT_DIR.exists():
        print(f"Tariff input folder not found: {TARIFF_INPUT_DIR}")
        return
    files = get_xlsx_files()
    if not files:
        print("No .xlsx files in 'input' folder.")
        return
    _ui_heading("Step 1 — Rate Card(s) from the 'input' folder")
    print(
        "These are the same Excel files that rate_to_json.py reads.\n"
        "You can pick one file, or merge several.\n"
    )
    print("Files:")
    for i, name in enumerate(files, 1):
        print(f"  {i}. {name}")
    choice = input("\nYour choice: ").strip()
    indices = parse_tariff_file_index_list(choice, len(files))
    if not indices:
        print("No valid file numbers.")
        return

    pairs: list[tuple[Path, str]] = []
    for fi in indices:
        filename = files[fi - 1]
        tariff_path = TARIFF_INPUT_DIR / filename
        wb = openpyxl.load_workbook(tariff_path, read_only=False, data_only=True)
        print(f"\nFile: {filename}")
        print("Sheets in this file:")
        for i, name in enumerate(wb.sheetnames, 1):
            print(f"  {i}. {name}")
        try:
            tab_idx = int(
                input(f"Which sheet number for this file? (1–{len(wb.sheetnames)}): ").strip()
            )
            sheet_name = wb.sheetnames[tab_idx - 1]
        except (ValueError, IndexError):
            print("Invalid sheet number.")
            return
        pairs.append((tariff_path, sheet_name))

    if len(pairs) == 1:
        clean_rows = export_tariff_sheet_rows(pairs[0][0], pairs[0][1])
        output_base = f"{pairs[0][0].stem}_{pairs[0][1]}"
    else:
        clean_rows = combine_tariff_sheet_rows(pairs)
        output_base = combined_tariff_output_basename(pairs)
        print("\nMerged tariff order (first file’s rows, then second, and so on):")
        for p, sn in pairs:
            print(f"  • {p.name}  →  sheet {sn!r}")

    _save_processing_tariff_jsons(output_base, clean_rows)
    print(
        f"\nSaved cleaned tariff JSON under:\n  {PROCESSING_DIR.resolve()}\n"
        f"Files: {output_base}.json and {output_base}_tariff_pipeline.json\n"
        "(Same row list as rate_to_json would produce — used to build the matrix.)"
    )
    if _DEBUG:
        _debug_step(1, "Tariff sheet → cleaned rows")
        for p, sn in pairs:
            _debug_line(f"Source: {p.name}, tab: {sn!r}", 0)
        _debug_line(f"output_base={output_base!r}", 0)
        _debug_line(f"Cleaned row count: {len(clean_rows)}", 0)
        if clean_rows:
            _debug_line(f"First row sample keys: {list(clean_rows[0].keys())[:12]}", 0)
        _debug_line(
            f"Tariff JSON snapshots → {PROCESSING_DIR.resolve()} "
            f"({output_base}.json + _tariff_pipeline.json)",
            0,
        )

    tables = _split_into_tables(clean_rows)
    if _DEBUG:
        _debug_step(2, "Split into tables (_split_into_tables)")
        _debug_line(f"Table blocks: {len(tables)}", 0)
        for ti, tb in enumerate(tables):
            blob = _table_title_blob(tb[0])[:120] if tb else ""
            _debug_line(f"  Block {ti + 1}: {len(tb)} rows, title blob: {blob!r}", 0)

    tables = resplit_tables_by_titles(tables)
    if _DEBUG:
        _debug_step(3, "Re-split (resplit_tables_by_titles) if one block had multiple titles")
        _debug_line(f"Table blocks now: {len(tables)}", 0)
        for ti, tb in enumerate(tables):
            blob = _table_title_blob(tb[0])[:120] if tb else ""
            _debug_line(f"  Block {ti + 1}: {len(tb)} rows, title: {blob!r}", 0)

    print(
        f"\nTariff data was split into {len(tables)} table block(s) "
        "(by title rows; large sheets may split further)."
    )
    origin_tb, return_tb = pick_origin_return_tables(tables)
    otitle = _table_title_blob(origin_tb[0])[:100] if origin_tb else ""
    rtitle = _table_title_blob(return_tb[0])[:100] if return_tb else ""
    if _DEBUG:
        _debug_step(4, "Assign origin vs return table (title contains 'return' / 'returns')")
        _debug_line(f"ORIGIN table first-row blob: {otitle!r}", 0)
        _debug_line(f"RETURN table first-row blob: {rtitle!r}", 0)
        _debug_line("If both are the same, only one physical table exists — return block reuses it.", 0)
    print("\nDefault pairing for outbound vs return lanes (from table titles):")
    print(f"  Outbound (not marked return): {otitle[:90] or '(empty)'}")
    print(f"  Return lane:                  {rtitle[:90] or '(empty)'}")

    # 2) Example rate card: previous rate folder
    if not PREVIOUS_RATE_DIR.exists():
        print(f"Folder not found: {PREVIOUS_RATE_DIR}")
        return
    prev_files = sorted(PREVIOUS_RATE_DIR.glob("*.xlsx"))
    if not prev_files:
        print("No .xlsx in 'previous rate' folder.")
        return
    _ui_heading("Step 2 — Example Rate Agreements")
    print(
        "Pick one workbook from the 'previous rate' folder. Its lanes and patterns drive the matrix columns.\n"
    )
    print("Files:")
    for i, f in enumerate(prev_files, 1):
        print(f"  {i}. {f.name}")
    try:
        p_idx = int(input(f"\nWhich file? (1–{len(prev_files)}): ").strip())
        if not (1 <= p_idx <= len(prev_files)):
            raise ValueError
        prev_path = prev_files[p_idx - 1]
    except (ValueError, IndexError):
        print("Invalid file number.")
        return

    rate_card_full = export_rate_card_full_json(prev_path)
    if not rate_card_full:
        print("No patterns extracted from example rate card.")
        return
    patterns = rate_card_full["patterns"]
    stable_keys = rate_card_full["pattern_analysis"]["stable_keys"]
    changing_keys = rate_card_full["pattern_analysis"]["changing_keys"]
    rates_table = rate_card_full["rates_table"]
    rate_card_costs = rate_card_full["costs"]
    PROCESSING_DIR.mkdir(parents=True, exist_ok=True)
    rc_snap = PROCESSING_DIR / f"{prev_path.stem}_rate_card.json"
    with open(rc_snap, "w", encoding="utf-8") as f:
        json.dump(rate_card_full, f, indent=2, ensure_ascii=False)
    changing_fill_keys = [k for k in (changing_keys or []) if k and k not in ("Lane #", "Lane")]
    if _DEBUG:
        print(f"\nFull rate card JSON written to:\n  {rc_snap}")
        print(f"\nFound {len(patterns)} lane pattern(s) in that rate card.")
        if rate_card_costs:
            print(
                f"Also found {len(rate_card_costs)} transport cost block(s) in the rate card "
                "(each becomes a horizontal block in the Excel output when applicable)."
            )
        if changing_fill_keys:
            print(
                "These pattern fields vary per row and will be filled from tariff columns where labels match:\n  "
                + ", ".join(changing_fill_keys)
            )
    if _DEBUG:
        _debug_step(5, "Rate card example → patterns (extract_patterns)")
        _debug_line(f"Example file: {prev_path.name}", 0)
        _debug_line(f"stable_keys (used as pattern column order): {stable_keys}", 0)
        _debug_line(f"changing_keys (vary per row in rate card; tariff fills these): {changing_fill_keys}", 0)
        for i, p in enumerate(patterns):
            _debug_line(f"  Pattern {i + 1} (index {i}): {p}", 0)
        _debug_line(
            "Patterns are enriched from rates_table where a field is constant for all rows matching "
            "the pattern (e.g. Origin Postal 73110, Return Destination Postal).",
            0,
        )
        _debug_line(f"rate_card transport_costs (horizontal blocks in Excel): {len(rate_card_costs)}", 0)

    origin_idx = 0
    return_idx: int | None = None
    flat_zero_idx: int | None = None
    origin_pat: dict[str, str] = patterns[0]
    return_pat: dict[str, str] | None = None
    flat_zero_pat: dict[str, str] | None = None
    flat_zero_base: str | None = None
    # At least two merged tariff xlsx + more than three rate-card patterns: assign a tariff table
    # block per pattern. Matches multi-sheet / multi-file tariffs with four lanes (e.g. Lahr + Lahr–CH).
    use_per_pattern_tables = len(patterns) > 3 and len(pairs) >= 2
    pattern_table_indices: list[int] | None = None
    enriched_all_patterns: list[dict] | None = None
    per_pattern_city_keys_from_card: list[frozenset[str]] | None = None

    if use_per_pattern_tables:
        per_pattern_city_keys_from_card = [_city_keys_declared_in_rate_card(p) for p in patterns]
        pattern_table_indices = _prompt_pattern_to_tariff_table_indices(patterns, tables)
        enriched_all_patterns = [
            _enrich_pattern_from_rates_table(p, rates_table, stable_keys) for p in patterns
        ]
        _ui_heading("Summary — each pattern linked to a tariff table")
        for i, p in enumerate(enriched_all_patterns):
            tb_i = (pattern_table_indices[i] + 1) if pattern_table_indices else 0
            role = _infer_pattern_lane_role(p)
            print(
                f"\n  Pattern {i + 1} → tariff table {tb_i} ({role} lane) · {_format_pattern_one_line_summary(p)}"
            )
            print(_format_pattern_detail_block(p, indent="    "))
    elif len(patterns) == 1:
        _ui_heading("Step 3 — Lane patterns")
        _print_pattern_catalog(
            patterns,
            intro="There is only one pattern. It will be used as the outbound (origin) lane. "
            "No separate return pattern.",
        )
        return_idx = None
        return_pat = None
    elif len(patterns) == 2:
        _ui_heading("Step 3 — Outbound vs return pattern")
        _print_pattern_catalog(
            patterns,
            intro="Define the initial pattern and return pattern.",
        )
        while True:
            raw = input("\nInitial pattern number — 1 or 2: ").strip()
            try:
                choice = int(raw)
                if choice in (1, 2):
                    break
            except ValueError:
                pass
            print("  Please type 1 or 2.")
        origin_idx = choice - 1
        return_idx = 1 - origin_idx
        origin_pat = patterns[origin_idx]
        return_pat = patterns[return_idx]
        print(
            f"\n  Confirmed: initial = pattern {origin_idx + 1}, return = pattern {return_idx + 1}."
        )
    elif len(patterns) == 3:
        _ui_heading("Step 3 — Outbound, flat-zero, and return")
        _print_pattern_catalog(
            patterns,
            intro="Three patterns were found. Assign each role once: outbound (origin), flat-zero, and return.\n"
            "Use three different numbers from 1 to 3.",
        )
        while True:
            try:
                o = int(input("\nOutbound (origin) pattern number (1–3): ").strip())
                fz = int(input("Flat-zero pattern number (1–3): ").strip())
                r = int(input("Return pattern number (1–3): ").strip())
                if len({o, fz, r}) == 3 and all(1 <= x <= 3 for x in (o, fz, r)):
                    origin_idx = o - 1
                    flat_zero_idx = fz - 1
                    return_idx = r - 1
                    break
            except ValueError:
                pass
            print("  Enter three different numbers, each between 1 and 3.")
        origin_pat = patterns[origin_idx]
        flat_zero_pat = patterns[flat_zero_idx]
        return_pat = patterns[return_idx]
        print(
            f"\n  Confirmed: outbound = pattern {origin_idx + 1}, flat-zero = pattern {flat_zero_idx + 1}, "
            f"return = pattern {return_idx + 1}."
        )
        while True:
            raw = input(
                "\nFor flat-zero rows, copy per-kg (p/unit) amounts from which tariff table? "
                "Type 'origin' for outbound or 'return' for return: "
            ).strip().lower()
            if raw in ("origin", "o", "1"):
                flat_zero_base = "origin"
                break
            if raw in ("return", "r", "2"):
                flat_zero_base = "return"
                break
            print("  Answer with 'origin' or 'return' (or o / r).")
    else:
        _ui_heading("Step 3 — Outbound vs return (many patterns)")
        _print_pattern_catalog(
            patterns,
            intro=f"There are {len(patterns)} patterns. Pick two different numbers: one outbound lane, one return lane.",
        )
        while True:
            try:
                o = int(input("\nOutbound (origin) pattern number: ").strip())
                r = int(input("Return pattern number: ").strip())
                if 1 <= o <= len(patterns) and 1 <= r <= len(patterns) and o != r:
                    origin_idx = o - 1
                    return_idx = r - 1
                    break
            except ValueError:
                pass
            print(
                f"  Enter two different numbers between 1 and {len(patterns)}."
            )
        origin_pat = patterns[origin_idx]
        return_pat = patterns[return_idx]
        print(
            f"\n  Confirmed: outbound = pattern {origin_idx + 1}, return = pattern {return_idx + 1}."
        )

    if not use_per_pattern_tables:
        origin_city_keys_card = _city_keys_declared_in_rate_card(origin_pat)
        return_city_keys_card = (
            _city_keys_declared_in_rate_card(return_pat) if return_pat is not None else frozenset()
        )
        flat_zero_city_keys_card = (
            _city_keys_declared_in_rate_card(flat_zero_pat) if flat_zero_pat is not None else frozenset()
        )
        origin_pat = _enrich_pattern_from_rates_table(origin_pat, rates_table, stable_keys)
        if flat_zero_pat is not None:
            flat_zero_pat = _enrich_pattern_from_rates_table(flat_zero_pat, rates_table, stable_keys)
        if return_pat is not None:
            return_pat = _enrich_pattern_from_rates_table(return_pat, rates_table, stable_keys)

    if use_per_pattern_tables:
        assert enriched_all_patterns is not None
        key_order = list(stable_keys) if stable_keys else []
        for k in _collect_pattern_key_order(enriched_all_patterns):
            if k not in key_order:
                key_order.append(k)
        for k in changing_fill_keys:
            if k not in key_order:
                key_order.append(k)
    else:
        key_order = list(stable_keys) if stable_keys else []
        for k in _collect_pattern_key_order(
            [origin_pat]
            + ([flat_zero_pat] if flat_zero_pat is not None else [])
            + ([return_pat] if return_pat else [])
        ):
            if k not in key_order:
                key_order.append(k)
        for k in changing_fill_keys:
            if k not in key_order:
                key_order.append(k)

    if _DEBUG:
        if use_per_pattern_tables:
            _debug_step(6, "Per-pattern tariff table blocks (merged xlsx + >3 patterns)")
            assert pattern_table_indices is not None and enriched_all_patterns is not None
            for i, ep in enumerate(enriched_all_patterns):
                _debug_line(
                    f"Pattern {i + 1} → table {pattern_table_indices[i] + 1} "
                    f"({_infer_pattern_lane_role(ep)}): {ep}",
                    0,
                )
        else:
            _debug_step(6, "Your origin / return / flat-zero pattern choice")
            _debug_line(f"origin_idx={origin_idx} → user-facing Pattern {origin_idx + 1}", 0)
            _debug_line(f"origin_pat after rates_table enrichment: {origin_pat}", 0)
            if flat_zero_idx is not None:
                _debug_line(f"flat_zero_idx={flat_zero_idx} → Pattern {flat_zero_idx + 1}", 0)
                _debug_line(f"flat_zero_pat after enrichment: {flat_zero_pat}", 0)
                _debug_line(f"flat_zero_base={flat_zero_base!r} (p/unit source tariff table)", 0)
            if return_idx is not None:
                _debug_line(f"return_idx={return_idx} → Pattern {return_idx + 1}", 0)
                _debug_line(f"return_pat after enrichment: {return_pat}", 0)
        _debug_line(f"key_order (pattern columns): {key_order}", 0)
        if not use_per_pattern_tables:
            _debug_line(
                "_pattern_role=origin | flat_zero | return; flat_zero uses Flat→0 and p/unit from zero-base.",
                0,
            )

    if use_per_pattern_tables:
        assert enriched_all_patterns is not None and pattern_table_indices is not None
        matrix_rows = _build_matrix_rows_multi(
            key_order,
            patterns=enriched_all_patterns,
            pattern_table_indices=pattern_table_indices,
            tables=tables,
            changing_keys=changing_fill_keys,
            pattern_city_keys_from_rate_card=per_pattern_city_keys_from_card,
        )
    else:
        matrix_rows = _build_matrix_rows(
            key_order,
            origin_pattern=origin_pat,
            return_pattern=return_pat,
            origin_table=origin_tb,
            return_table=return_tb if return_pat is not None else None,
            origin_pattern_idx=origin_idx,
            return_pattern_idx=return_idx,
            changing_keys=changing_fill_keys,
            flat_zero_pattern=flat_zero_pat,
            flat_zero_pattern_idx=flat_zero_idx,
            flat_zero_base=flat_zero_base,
            origin_city_keys_from_rate_card=origin_city_keys_card,
            return_city_keys_from_rate_card=return_city_keys_card,
            flat_zero_city_keys_from_rate_card=flat_zero_city_keys_card,
        )
    if _DEBUG and rate_card_costs:
        print(
            f"\nBuilding the matrix: {len(rate_card_costs)} transport cost block(s) from the rate card "
            "will expand into separate horizontal blocks in Excel (currency + bands per cost)."
        )
    if _DEBUG and matrix_rows:
        _debug_step(7, "DataFrame column order (why you might not 'see' a pattern column)")
        _debug_line(
            "Shipment columns follow the previous rate card rates_table column order (no _pattern_* in export).",
            0,
        )
        _debug_line(f"First row keys (sample): {list(matrix_rows[0].keys())[:25]}", 0)

    if not matrix_rows:
        print(
            "\nNo data rows were written to the matrix. Things to check:\n"
            "  • The tariff tab should contain rate tables (titles like 'New Freight Rate' / 'Rates RDE', "
            "or one continuous block).\n"
            "  • Rows should look like weight bands with amounts (not only headers or notes).\n"
            "  • Try another sheet if this one has no freight table.\n"
        )
        print(
            f"  (Internal: {len(clean_rows)} cleaned tariff rows, {len(tables)} table block(s).)"
        )
        for ti, tb in enumerate(tables):
            lbls, dr = _parse_header_and_data(tb)
            print(f"    Table {ti + 1}: {len(tb)} rows in block, {len(dr)} data rows detected.")
            if tb and not dr:
                k = _primary_column_key(tb[0])
                print(f"      First row keys sample: {list(tb[0].keys())[:8]} primary_col={k!r}")
        return

    # Column order: rate-card lane columns (order + presence), tariff lane_misc, then rate bands.
    # Internal keys (_pattern_*, _tariff_table_block) are omitted from Excel/JSON output.
    reserved = _EXCEL_OUTPUT_INTERNAL_DROP
    pattern_cols = set(_all_pattern_output_keys(key_order))
    extra_cols = []
    for r in matrix_rows:
        for k in r:
            if k not in pattern_cols and k not in reserved:
                if k not in extra_cols:
                    extra_cols.append(k)
    extra_cols = sort_extra_tariff_columns(extra_cols)
    _, rate_block_cols = build_rate_block_column_order(extra_cols)

    df = pd.DataFrame(matrix_rows)
    df = _apply_chosen_tariff_sheet_to_matrix(df, rates_table, pairs)
    shipment_prefix = _excel_shipment_column_prefix(df, rates_table, key_order)
    # Tariff-only lane columns (Zip 2, ColumnN, …) are omitted — export matches previous rate card + rate bands.
    ordered = shipment_prefix + rate_block_cols
    df = df.reindex(columns=[c for c in ordered if c in df.columns] + [c for c in df.columns if c not in ordered])
    _n_before = len(df.columns)
    df = drop_all_empty_columns(df)
    _n_drop = _n_before - len(df.columns)
    if _DEBUG and _n_drop:
        print(f"\nRemoved {_n_drop} column(s) that were empty in every row.")

    rate_block_cols = [c for c in rate_block_cols if c in df.columns]
    df, rate_block_cols = normalize_currency_column_and_strip_band_amounts(df, rate_block_cols)
    shipment_prefix = _excel_shipment_column_prefix(df, rates_table, key_order)
    ordered = shipment_prefix + rate_block_cols
    df = df.reindex(columns=[c for c in ordered if c in df.columns] + [c for c in df.columns if c not in ordered])
    _curr_keep = frozenset(c for c in df.columns if _is_currency_column_name(str(c)))
    df = drop_all_empty_columns(df, keep_columns=_curr_keep if _curr_keep else None)

    if use_per_pattern_tables:
        measure_by_label: dict[str, str] = {}
        for tb in tables:
            measure_by_label.update(extract_rate_measure_labels_by_display_name(tb))
    else:
        measure_by_label = {
            **extract_rate_measure_labels_by_display_name(return_tb or []),
            **extract_rate_measure_labels_by_display_name(origin_tb),
        }
    rate_column_measures = _measure_list_for_rate_block_columns(rate_block_cols, measure_by_label)
    _ui_heading("Optional - manual '>=' FLAT columns")
    if _prompt_add_synthetic_geq_columns():
        fill_from_ftl = False
        if any(_is_ftl_tariff_column(c) for c in rate_block_cols):
            fill_from_ftl = _prompt_synthetic_geq_fill_mode() == "ftl"
        df, rate_block_cols, rate_column_measures = add_synthetic_geq_flat_after_last_p_unit(
            df,
            rate_block_cols,
            rate_column_measures,
            fill_synthetic_from_ftl=fill_from_ftl,
        )
    shipment_prefix = _excel_shipment_column_prefix(df, rates_table, key_order)
    ordered = shipment_prefix + rate_block_cols
    df = df.reindex(columns=[c for c in ordered if c in df.columns] + [c for c in df.columns if c not in ordered])

    shipment_cols = [c for c in ordered if c in df.columns and c not in rate_block_cols]
    use_wide = bool(rate_card_costs) and bool(rate_block_cols)
    sorted_costs = sort_transport_costs_non_ftl_first(
        [c for c in (rate_card_costs or []) if isinstance(c, dict)]
    )
    rate_block_cols_by_cost = [
        rate_block_cols_for_transport_cost(c, rate_block_cols) for c in sorted_costs
    ]
    rate_column_measures_by_cost = [
        _measure_list_for_rate_block_columns(rb, measure_by_label) for rb in rate_block_cols_by_cost
    ]
    if use_wide:
        df_out, col_order_final = widen_matrix_for_horizontal_cost_blocks(
            df,
            sorted_costs,
            shipment_cols,
            rate_block_cols,
            rate_block_cols_by_cost=rate_block_cols_by_cost,
        )
    else:
        df_out = df
        col_order_final = [c for c in ordered if c in df.columns]

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    out_name = f"{output_base}_rate_matrix.xlsx"
    out_path = OUTPUT_DIR / out_name
    col_order_final = [c for c in col_order_final if c in df_out.columns] + [
        c for c in df_out.columns if c not in col_order_final
    ]
    write_rate_matrix_excel_advanced_export_layout(
        df_out,
        out_path,
        column_order=col_order_final,
        transport_costs=sorted_costs if use_wide else (rate_card_costs if rate_card_costs else None),
        shipment_cols=shipment_cols,
        rate_block_cols=rate_block_cols,
        wide_format=use_wide,
        rate_column_measures=rate_column_measures,
        rate_block_cols_by_cost=rate_block_cols_by_cost if use_wide else None,
        rate_column_measures_by_cost=rate_column_measures_by_cost if use_wide else None,
    )
    layout_note = " — wide layout with one horizontal block per transport cost" if use_wide else ""
    _ui_heading("Output files")
    print(
        f"Excel matrix ({len(df_out)} rows × {len(df_out.columns)} columns){layout_note}:\n  {out_path}"
    )

    json_path = out_path.with_suffix(".json")
    payload = {
        "transport_costs": sorted_costs if use_wide else rate_card_costs,
        "shipment_columns": [c for c in shipment_cols if c in df_out.columns],
        "rate_block_columns": rate_block_cols,
        "rate_block_columns_by_cost": rate_block_cols_by_cost if use_wide else None,
        "rate_column_measures": {
            col: rate_column_measures[i]
            for i, col in enumerate(rate_block_cols)
            if i < len(rate_column_measures) and str(rate_column_measures[i] or "").strip()
        },
        "matrix_wide_key_note": (
            "Each cost i uses rate_block_columns_by_cost[i]; matrix keys __bk{i}_c{j}__ for "
            "j = 0 .. len(rate_block_columns_by_cost[i])-1. Non-FTL costs omit the FTL tariff column; "
            "FTL-named costs use Currency + FTL only when the tariff has an FTL column."
        ),
        "matrix": json.loads(df_out.to_json(orient="records", date_format="iso", default_handler=str)),
    }
    if (
        use_per_pattern_tables
        and pattern_table_indices is not None
        and enriched_all_patterns is not None
    ):
        payload["patterns_with_tariff_table_block"] = [
            {
                **dict(enriched_all_patterns[i]),
                "tariff_table_block": pattern_table_indices[i] + 1,
                "inferred_lane_role": _infer_pattern_lane_role(enriched_all_patterns[i]),
            }
            for i in range(len(enriched_all_patterns))
        ]
    for dest in (OUTPUT_DIR, PROCESSING_DIR):
        dest.mkdir(parents=True, exist_ok=True)
        jp = dest / json_path.name
        with open(jp, "w", encoding="utf-8") as f:
            json.dump(payload, f, indent=2, ensure_ascii=False)
    print(f"\nJSON (same data + metadata):\n  {json_path}")
    print(f"Copy for processing folder:\n  {PROCESSING_DIR / json_path.name}")

    # For Colab / run_workflow: rename outputs and archive inputs after the run.
    try:
        wf_meta = {
            "prev_rate_file": str(prev_path.resolve()),
            "prev_rate_stem": prev_path.stem,
            "output_matrix_basename": out_path.name,
            "tariff_input_files": [str(p.resolve()) for p, _ in pairs],
        }
        with open(PROCESSING_DIR / "last_run_workflow.json", "w", encoding="utf-8") as f:
            json.dump(wf_meta, f, indent=2, ensure_ascii=False)
    except Exception:
        pass


if __name__ == "__main__":
    main()



